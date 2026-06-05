"""
build_site.py — Render the live portfolio into the static GitHub Pages site (v2).

Outputs (into docs/, which GitHub Pages serves):
  * docs/data.json     : full portfolio snapshot (machine-readable)
  * docs/history.json  : appended weekly time-series for the trend/performance views
  * docs/index.html    : branded, self-contained client dashboard (v2)

v2 design: clean institutional light, tabbed sections (Overview / Allocation /
Holdings / Performance / Risk), with concentration, currency and risk-posture
analytics. The page embeds the snapshot inline so it renders as a local file and
also re-fetches data.json so a hosted copy always shows the latest. Performance
and week-over-week visuals use the REAL accumulating Weekly Snapshot history and
degrade gracefully to a "history is building" state until a second point exists.
"""
from __future__ import annotations

import argparse
import datetime as _dt
import json
from pathlib import Path

from portfolio import load_portfolio, WORKBOOK_DEFAULT

ROOT = Path(__file__).resolve().parent.parent
DOCS = ROOT / "docs"
CONFIG = Path(__file__).resolve().parent / "config.json"


def update_history(snapshot: dict, asof: str) -> list:
    hist_path = DOCS / "history.json"
    history = []
    if hist_path.exists():
        try:
            history = json.loads(hist_path.read_text())
        except json.JSONDecodeError:
            history = []
    row = {
        "date": asof,
        "total_value": round(snapshot["total_value"]),
        "total_invested": round(snapshot["total_invested"]),
        "total_pnl": round(snapshot["total_pnl"]),
        "pnl_pct": round(snapshot["total_pnl_pct"], 4),
        "mf_value": round(snapshot["mf_value"]),
        "eq_value": round(snapshot["eq_value"]),
        "crypto_value": round(snapshot["crypto_value"]),
        "fx_rate": snapshot["fx_rate"],
    }
    history = [h for h in history if h.get("date") != asof]  # replace same-day
    history.append(row)
    history.sort(key=lambda h: h["date"])
    hist_path.write_text(json.dumps(history, indent=2))
    return history


def read_flows(wb) -> list:
    """External cash flows from the 'Cash Flows' sheet (date, amount ฿, type, note).
    Positive = money in (contribution/dividend), negative = money out (withdrawal).
    Used by the Overview daily card to attribute real flows vs market/FX movement —
    do NOT infer flows from changes in cost basis, which drift daily with FX."""
    flows = []
    if "Cash Flows" not in wb.sheetnames:
        return flows
    ws = wb["Cash Flows"]
    for r in range(5, ws.max_row + 1):
        dt_, amt = ws.cell(r, 1).value, ws.cell(r, 2).value
        if dt_ is None or amt in (None, ""):
            continue
        if isinstance(dt_, _dt.datetime):
            dt_ = dt_.date().isoformat()
        elif isinstance(dt_, _dt.date):
            dt_ = dt_.isoformat()
        else:
            dt_ = str(dt_)[:10]
        try:
            amt = float(amt)
        except (TypeError, ValueError):
            continue
        flows.append({"date": dt_, "amount": round(amt),
                      "type": ws.cell(r, 3).value or "", "note": ws.cell(r, 4).value or ""})
    return flows


def build(workbook=WORKBOOK_DEFAULT, generated_at=None):
    DOCS.mkdir(exist_ok=True)
    cfg = json.loads(CONFIG.read_text())
    p = load_portfolio(workbook).to_dict()
    history = update_history(p, p["as_of"])

    generated_at = generated_at or _dt.datetime.now().strftime("%Y-%m-%d %H:%M")

    # institutional performance metrics (benchmark-relative, TWR/IRR, risk stats)
    flows = []
    try:
        import openpyxl
        import performance
        wb = openpyxl.load_workbook(workbook, data_only=False)
        perf = performance.compute(p, wb, generated_at)
        flows = read_flows(wb)
    except Exception as e:  # never let perf break the core dashboard
        perf = {"error": str(e), "history_points": len(history)}

    try:
        import tax as _tax
        taxobj = _tax.compute(p, wb)
    except Exception as e:
        taxobj = {"error": str(e)}

    payload = {"config": cfg, "portfolio": p, "history": history, "flows": flows,
               "performance": perf, "tax": taxobj, "generated_at": generated_at}
    (DOCS / "data.json").write_text(json.dumps(payload, indent=2, default=str))

    html = HTML_TEMPLATE.replace("/*__DATA__*/null", json.dumps(payload, default=str))
    (DOCS / "index.html").write_text(html)
    return payload


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover"/>
<title>Private Banking Summary</title>
<meta name="theme-color" content="#0b3d2e"/>
<link rel="manifest" href="manifest.webmanifest"/>
<link rel="apple-touch-icon" href="apple-touch-icon.png"/>
<meta name="apple-mobile-web-app-capable" content="yes"/>
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent"/>
<meta name="apple-mobile-web-app-title" content="TH Wealth"/>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
  :root{
    --bg:#f6f7f8; --surface:#ffffff; --surface-2:#fbfcfc;
    --border:#e6e8eb; --border-2:#d7dbe0;
    --ink:#14181c; --muted:#69727b; --faint:#98a0a8;
    --accent:#0b3d2e; --accent-2:#15634a; --accent-tint:#eef3f1;
    --pos:#0f7a44; --pos-bg:#e8f4ee; --neg:#c0392b; --neg-bg:#fbecea;
    --gold:#b9975b;
    --r:10px; --shadow:0 1px 2px rgba(16,24,32,.04),0 1px 1px rgba(16,24,32,.03);
  }
  *{box-sizing:border-box}
  html,body{margin:0;-webkit-text-size-adjust:100%;text-size-adjust:100%}
  body{font-family:-apple-system,BlinkMacSystemFont,"Inter","Segoe UI",Roboto,Helvetica,Arial,sans-serif;
       background:var(--bg);color:var(--ink);-webkit-font-smoothing:antialiased;font-size:14px;line-height:1.45;overflow-x:hidden}
  .num{font-variant-numeric:tabular-nums;font-feature-settings:"tnum" 1}
  .pos{color:var(--pos)} .neg{color:var(--neg)}
  a{color:inherit}

  .topbar{position:sticky;top:0;z-index:40;background:rgba(255,255,255,.92);backdrop-filter:saturate(1.4) blur(8px);
    border-bottom:1px solid var(--border)}
  .topbar .in{max-width:1240px;margin:0 auto;padding:11px 22px;display:flex;align-items:center;gap:18px}
  .wordmark{display:flex;align-items:center;gap:10px;font-weight:680;letter-spacing:-.01em}
  .logo{width:26px;height:26px;border-radius:7px;background:linear-gradient(135deg,var(--accent),var(--accent-2));
    display:grid;place-items:center;color:#fff;font-size:12px;font-weight:700}
  .who{color:var(--muted);font-size:12.5px;padding-left:14px;border-left:1px solid var(--border);margin-left:2px}
  .who b{color:var(--ink);font-weight:620}
  .spacer{flex:1}
  .meta{display:flex;gap:18px;align-items:center;font-size:12.5px;color:var(--muted)}
  .meta .v{color:var(--ink);font-weight:600}
  .tag{font-size:10.5px;letter-spacing:.06em;text-transform:uppercase;color:var(--gold);
    border:1px solid #e7dcc3;background:#fbf7ee;padding:3px 8px;border-radius:20px}

  .tabs{position:sticky;top:49px;z-index:30;background:var(--bg);border-bottom:1px solid var(--border)}
  .tabs .in{max-width:1240px;margin:0 auto;padding:0 22px;display:flex;gap:4px;overflow-x:auto}
  .tab{appearance:none;border:0;background:none;font:inherit;font-size:13.5px;color:var(--muted);
    padding:13px 14px 11px;cursor:pointer;border-bottom:2px solid transparent;white-space:nowrap;font-weight:550}
  .tab:hover{color:var(--ink)}
  .tab.active{color:var(--accent);border-bottom-color:var(--accent);font-weight:640}

  /* mobile bottom navigation (thumb-reachable) */
  .bottomnav{display:none}
  .cardlist{display:none}
  @media(max-width:640px){
    .tabs{display:none}
    .bottomnav{display:flex;position:fixed;left:0;right:0;bottom:0;z-index:60;
      background:rgba(255,255,255,.97);backdrop-filter:saturate(1.4) blur(12px);
      border-top:1px solid var(--border);padding-bottom:env(safe-area-inset-bottom)}
    .bottomnav button{flex:1;border:0;background:none;font:inherit;cursor:pointer;color:var(--muted);
      display:flex;flex-direction:column;align-items:center;gap:3px;padding:8px 1px 7px;
      font-size:9.5px;font-weight:600;letter-spacing:.01em;-webkit-tap-highlight-color:transparent}
    .bottomnav button.active{color:var(--accent)}
    .bottomnav svg{width:21px;height:21px;stroke:currentColor;fill:none;stroke-width:1.7;stroke-linecap:round;stroke-linejoin:round}
    .bottomnav button.active svg{stroke-width:2.2}
    .wrap{padding-bottom:74px}
    /* swap wide scroll-tables for tappable cards on Holdings & Tax */
    .tablewrap{display:none}
    .cardlist{display:block}
  }
  .lcard{padding:12px 2px;border-bottom:1px solid var(--border);cursor:pointer;-webkit-tap-highlight-color:transparent}
  .lcard-top{display:flex;justify-content:space-between;gap:12px;align-items:flex-start}
  .lcard-nm{font-weight:600;font-size:14px;min-width:0;overflow-wrap:anywhere;line-height:1.25}
  .lcard-nm .sub{font-weight:400}
  .lcard-rt{text-align:right;white-space:nowrap;flex-shrink:0}
  .lcard-rt .v{font-size:13.5px;font-weight:600}
  .lcard-rt .r{font-size:12px;font-weight:600;display:block;margin-top:1px}
  .lcard-more{display:none;margin-top:10px;grid-template-columns:1fr 1fr;gap:6px 14px;font-size:11.5px;color:var(--muted)}
  .lcard.open .lcard-more{display:grid}
  .lcard-more div{display:flex;justify-content:space-between;gap:8px;border-bottom:1px dotted var(--border);padding-bottom:3px}
  .lcard-more b{color:var(--ink);font-weight:600;font-variant-numeric:tabular-nums}
  .lcard .caret{color:var(--faint);font-size:10px;margin-left:6px;display:inline-block;transition:transform .15s}
  .lcard.open .caret{transform:rotate(180deg)}

  .wrap{max-width:1240px;margin:0 auto;padding:22px}
  section[hidden]{display:none}
  .row{display:grid;gap:16px}
  .g4{grid-template-columns:repeat(4,1fr)} .g3{grid-template-columns:repeat(3,1fr)}
  .g2{grid-template-columns:repeat(2,1fr)} .g23{grid-template-columns:1.4fr 1fr}
  .g6{grid-template-columns:repeat(6,1fr)} .g155{grid-template-columns:1.55fr 1fr}
  .card{background:var(--surface);border:1px solid var(--border);border-radius:var(--r);box-shadow:var(--shadow);min-width:0}
  .card.pad{padding:18px}
  .card h3{margin:0 0 2px;font-size:13px;font-weight:640;letter-spacing:.01em}
  .card .hint{color:var(--faint);font-size:11.5px;margin-bottom:12px}
  .head{display:flex;justify-content:space-between;align-items:baseline;margin-bottom:12px}
  .head h3{margin:0}
  .mini{font-size:11.5px;color:var(--faint)}

  .kpi{padding:16px 16px 12px;position:relative;overflow:hidden}
  .kpi .label{font-size:11px;letter-spacing:.06em;text-transform:uppercase;color:var(--muted)}
  .kpi .val{font-size:25px;font-weight:680;letter-spacing:-.02em;margin-top:7px}
  .delta{display:inline-flex;align-items:center;gap:4px;font-size:11.5px;font-weight:600;margin-top:8px;
    padding:2px 7px;border-radius:20px}
  .delta.up{color:var(--pos);background:var(--pos-bg)} .delta.down{color:var(--neg);background:var(--neg-bg)}
  .spark{position:absolute;right:10px;bottom:10px;width:74px;height:30px;opacity:.9}

  /* Today's Movement card (Overview): day-over-day change + asset-bucket breakdown */
  .movcard{padding:16px 18px}
  .movmain{display:flex;justify-content:space-between;align-items:flex-start;gap:16px 24px;flex-wrap:wrap}
  .movbig{font-size:29px;font-weight:700;letter-spacing:-.02em;margin-top:5px;display:flex;align-items:baseline;gap:10px;flex-wrap:wrap}
  .movpct{font-size:13px;font-weight:600;padding:2px 8px;border-radius:20px;letter-spacing:0}
  .movpct.up{color:var(--pos);background:var(--pos-bg)} .movpct.down{color:var(--neg);background:var(--neg-bg)}
  .movbreak{display:flex;flex-wrap:wrap;gap:7px;align-items:center;max-width:100%}
  .movchip{font-size:11.5px;font-weight:600;padding:3px 9px;border-radius:7px;border:1px solid var(--border);white-space:nowrap}
  .movchip.up{color:var(--pos)} .movchip.down{color:var(--neg)} .movchip.flat{color:var(--muted)}

  .barstack{height:22px;border-radius:6px;overflow:hidden;display:flex;background:var(--border)}
  .barstack i{display:block;height:100%}
  .legend{display:flex;flex-wrap:wrap;gap:8px 16px;margin-top:14px;font-size:12px;color:var(--muted)}
  .legend div{display:flex;align-items:center;gap:7px}
  .dot{width:9px;height:9px;border-radius:3px;display:inline-block}
  table{width:100%;border-collapse:collapse}
  th,td{text-align:right;padding:9px 10px;font-size:12.5px;white-space:nowrap}
  th:first-child,td:first-child{text-align:left}
  /* Holding/lot name: fixed ~34ch column so names wrap to a 2nd line only past
     ~34 chars (~95% stay 1-line) and don't get squeezed narrower on mobile */
  #holdTable td:first-child,#taxLots td:first-child{white-space:normal;min-width:34ch;max-width:34ch;line-height:1.32}
  /* Holdings desktop: cap Class & Geo (and let them wrap) so the Priced column
     stays visible without horizontal scrolling at typical desktop widths */
  #holdTable th:nth-child(3),#holdTable td:nth-child(3){max-width:78px}
  #holdTable th:nth-child(4),#holdTable td:nth-child(4){max-width:66px}
  #holdTable td:nth-child(3),#holdTable td:nth-child(4){white-space:normal;line-height:1.3}  /* wrap multi-word values at spaces, keep single words (Vietnam, Thailand) intact */
  #holdTable th,#holdTable td{padding-left:6px;padding-right:6px}  /* tighter gutters so all 12 cols fit */
  /* Long names elsewhere wrap instead of overflowing (top movers / largest / concentration) */
  #ovHold td:first-child{white-space:normal;overflow-wrap:anywhere;line-height:1.3;max-width:26ch}
  .conc .nm{overflow-wrap:anywhere;line-height:1.25}
  thead th{font-size:10.5px;letter-spacing:.05em;text-transform:uppercase;color:var(--muted);font-weight:600;
    border-bottom:1px solid var(--border-2);cursor:pointer;user-select:none}
  tbody td{border-bottom:1px solid var(--border)}
  tbody tr:hover{background:var(--surface-2)}
  .name{font-weight:600}
  .sub{color:var(--faint);font-size:11px}
  /* "Priced" column: date of the latest price/NAV fetch. fresh=today, stale=>1wk old */
  .pdate{font-size:11px;color:var(--muted);white-space:nowrap}
  .pdate.fresh{color:var(--ink);font-weight:600}
  .pdate.stale{color:#b07a16}
  .chip{display:inline-block;font-size:10.5px;color:var(--muted);background:#f1f3f5;border:1px solid var(--border);
    padding:1px 7px;border-radius:5px}
  .wbar{height:6px;border-radius:4px;background:var(--accent-tint);overflow:hidden;min-width:40px}
  .wbar>i{display:block;height:100%;background:var(--accent-2)}
  .scroll{overflow-x:auto;-webkit-overflow-scrolling:touch;min-width:0}
  .row>*{min-width:0}
  .chartbox{position:relative;height:260px}

  .toolbar{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:14px}
  .search{flex:1;min-width:180px;border:1px solid var(--border-2);border-radius:8px;padding:8px 11px;font:inherit;font-size:13px;background:#fff}
  .seg{display:inline-flex;border:1px solid var(--border-2);border-radius:8px;overflow:hidden}
  .seg button{appearance:none;border:0;background:#fff;font:inherit;font-size:12.5px;color:var(--muted);padding:8px 12px;cursor:pointer;border-left:1px solid var(--border)}
  .seg button:first-child{border-left:0}
  .seg button.active{background:var(--accent);color:#fff}

  .stat{padding:16px}
  .stat .k{font-size:11px;letter-spacing:.05em;text-transform:uppercase;color:var(--muted)}
  .stat .v{font-size:23px;font-weight:680;margin-top:6px;letter-spacing:-.01em}
  .stat .d{font-size:11.5px;color:var(--faint);margin-top:3px}
  .conc ul{margin:0;padding:0}
  .conc li{display:grid;grid-template-columns:130px 1fr 56px;gap:10px;align-items:center;padding:6px 0;list-style:none}
  .conc .nm{font-size:12.5px;font-weight:560}
  .ribbon{display:flex;height:30px;border-radius:7px;overflow:hidden;border:1px solid var(--border)}
  .ribbon i{display:flex;align-items:center;justify-content:center;font-size:10.5px;color:#fff;font-weight:600;min-width:0}

  .note{font-size:11.5px;color:var(--faint);margin-top:10px}
  .empty{padding:26px;text-align:center;color:var(--muted);font-size:13px}
  .empty .pillnote{display:inline-block;background:#fff;border:1px dashed var(--border-2);color:var(--muted);
    font-size:12px;border-radius:20px;padding:6px 14px;margin-top:8px}
  footer{max-width:1240px;margin:8px auto 40px;padding:0 22px;color:var(--faint);font-size:11px;line-height:1.6}

  @media(max-width:980px){.g4{grid-template-columns:repeat(2,1fr)}.g3,.g2,.g23,.g155{grid-template-columns:1fr}.g6{grid-template-columns:repeat(3,1fr)}}
  @media(max-width:560px){.g6{grid-template-columns:repeat(2,1fr)}}
  @media(max-width:560px){
    .topbar{position:static}
    .topbar .in{flex-wrap:wrap;gap:8px 14px;padding:10px 14px}
    .who{padding-left:0;border-left:0;margin-left:0}
    .meta{gap:12px;font-size:12px} .tag{display:none}
    .tabs{top:0;position:sticky}
    .wrap{padding:16px 14px}
    .g4{grid-template-columns:1fr 1fr}
    .kpi .val{font-size:21px} .spark{display:none}
  }
</style>
</head>
<body>
<div class="topbar"><div class="in">
  <div class="wordmark"><span class="logo" id="logo">TH</span> <span id="firm">TH Investment</span></div>
  <div class="who"><b id="client">Private Client Portfolio</b> · <span id="ref"></span></div>
  <div class="spacer"></div>
  <div class="meta">
    <div>As of <span class="v" id="asof">—</span></div>
    <div>USD/THB <span class="v num" id="fx">—</span></div>
    <span class="tag">Confidential</span>
  </div>
</div></div>

<div class="tabs"><div class="in" id="tabnav"></div></div>

<div class="wrap">
  <section data-tab="Overview">
    <div class="card movcard" id="dailyMove" style="margin-bottom:16px"></div>
    <div class="row g4" id="kpis" style="margin-bottom:16px"></div>
    <div class="row g23" style="margin-bottom:16px">
      <div class="card pad">
        <div class="head"><h3>Asset Allocation</h3><span class="mini" id="acTotal"></span></div>
        <div class="barstack" id="acBar"></div>
        <div class="legend" id="acLegend"></div>
      </div>
      <div class="card pad">
        <div class="head"><h3>Top Movers</h3><span class="mini">by return</span></div>
        <div id="movers"></div>
      </div>
    </div>
    <div class="row g23">
      <div class="card pad">
        <div class="head"><h3>Performance</h3><span class="mini">value vs invested</span></div>
        <div id="ovPerfWrap" class="chartbox" style="height:230px"><canvas id="ovPerf"></canvas></div>
      </div>
      <div class="card pad">
        <div class="head"><h3>Largest Holdings</h3><span class="mini">top 6</span></div>
        <div class="scroll"><table id="ovHold"></table></div>
      </div>
    </div>
  </section>

  <section data-tab="Allocation" hidden>
    <div class="row g3" style="margin-bottom:16px">
      <div class="card pad"><h3>By Asset Class</h3><div class="hint">share of portfolio</div><div class="chartbox"><canvas id="acDonut"></canvas></div></div>
      <div class="card pad"><h3>By Geography</h3><div class="hint">look-through region</div><div id="geoBars"></div></div>
      <div class="card pad"><h3>By Tax Wrapper</h3><div class="hint">Thai vehicle type</div><div class="chartbox"><canvas id="taxDonut"></canvas></div></div>
    </div>
    <div class="row g2">
      <div class="card pad"><div class="head"><h3>Asset Class Detail</h3></div><div class="scroll"><table id="acTable"></table></div></div>
      <div class="card pad"><div class="head"><h3>Theme / Sector Exposure</h3></div><div id="themeBars"></div></div>
    </div>
  </section>

  <section data-tab="Holdings" hidden>
    <div class="card pad">
      <div class="toolbar">
        <input class="search" id="search" placeholder="Search holdings, class, geography…"/>
        <div class="seg" id="typeSeg"></div>
        <span class="mini" id="holdCount"></span>
      </div>
      <div class="scroll tablewrap"><table id="holdTable"></table></div>
      <div class="cardlist" id="holdCards"></div>
      <div class="hint" style="margin-top:12px"><b>Priced</b> = date the latest price/NAV was fetched —
        <b class="pdate fresh" style="font-size:inherit">bold</b> updated today,
        <span class="pdate stale" style="font-size:inherit">amber</span> over a week old (carried forward).</div>
    </div>
  </section>

  <section data-tab="Performance" hidden>
    <div class="row g6" style="margin-bottom:16px" id="perfStats"></div>
    <div class="card pad" style="margin-bottom:16px">
      <div class="head"><h3>Returns by Period</h3><span class="mini" id="perfBmkLbl"></span></div>
      <div class="scroll"><table id="perfPeriod"></table></div>
    </div>
    <div class="row g155" style="margin-bottom:16px">
      <div class="card pad">
        <div class="head"><h3>Growth of ฿100 — Portfolio vs Benchmark</h3>
          <div class="seg" id="perfRange"><button data-r="365">1Y</button><button data-r="1095">3Y</button><button class="active" data-r="1825">5Y</button><button data-r="99999">SI</button></div></div>
        <div id="perfGrowthWrap" class="chartbox" style="height:280px"><canvas id="perfGrowth"></canvas></div>
        <div class="legend"><div><span class="dot" style="background:var(--accent)"></span>Portfolio</div><div><span class="dot" style="background:var(--gold)"></span><span id="perfBmkName">Benchmark</span></div></div>
      </div>
      <div class="card pad">
        <div class="head"><h3>Risk &amp; Return</h3><span class="mini">annualized</span></div>
        <table id="perfRisk"></table>
        <div class="head" style="margin-top:16px"><h3>Drawdown</h3><span class="mini">peak-to-trough</span></div>
        <div id="perfDDWrap" class="chartbox" style="height:140px"><canvas id="perfDD"></canvas></div>
      </div>
    </div>
    <div class="card pad">
      <div class="head"><h3>Calendar-Year Returns</h3><span class="mini">portfolio vs benchmark</span></div>
      <div class="chartbox" style="height:240px"><canvas id="perfCal"></canvas></div>
    </div>
    <div class="note" id="perfNote"></div>
  </section>

  <section data-tab="Risk" hidden>
    <div class="row g4" id="riskStats" style="margin-bottom:16px"></div>
    <div class="row g2" style="margin-bottom:16px">
      <div class="card pad">
        <div class="head"><h3>Single-Name Concentration</h3><span class="mini">top 10 weights</span></div>
        <div class="conc"><ul id="concList"></ul></div>
      </div>
      <div class="card pad">
        <div class="head"><h3>Currency Exposure</h3><span class="mini">pre-hedge</span></div>
        <div class="barstack" id="fxBar" style="height:26px"></div>
        <div class="legend" id="fxLegend"></div>
        <div class="note">USD exposure is unhedged and revalues with the USD/THB rate.</div>
      </div>
    </div>
    <div class="card pad">
      <div class="head"><h3>Risk Posture</h3><span class="mini">by asset-class risk band</span></div>
      <div class="ribbon" id="riskRibbon"></div>
      <div class="legend" id="riskLegend"></div>
    </div>
  </section>

  <section data-tab="Tax &amp; Lots" hidden>
    <div class="row g6" style="margin-bottom:16px" id="taxStats"></div>
    <div class="row g155" style="margin-bottom:16px">
      <div class="card pad">
        <div class="head"><h3>Lock-up &amp; Maturity Schedule</h3><span class="mini">฿ unlocking by year (RMF / SSF / Thai ESG / LTF)</span></div>
        <div class="chartbox" style="height:240px"><canvas id="taxMaturity"></canvas></div>
        <div class="note" id="taxAvailNote"></div>
      </div>
      <div class="card pad">
        <div class="head"><h3>By Tax Wrapper</h3></div>
        <div class="scroll"><table id="taxWrapper"></table></div>
      </div>
    </div>
    <div class="card pad">
      <div class="head"><h3>Lots</h3>
        <div class="toolbar" style="margin:0">
          <input class="search" id="taxSearch" placeholder="Search lots…" style="min-width:150px"/>
          <div class="seg" id="taxSeg"></div>
          <span class="mini" id="taxCount"></span>
        </div>
      </div>
      <div class="scroll tablewrap"><table id="taxLots"></table></div>
      <div class="cardlist" id="taxCards"></div>
    </div>
    <div class="note" id="taxNote"></div>
  </section>
</div>

<nav class="bottomnav" id="bottomnav"></nav>
<footer id="foot"></footer>

<script>
const EMBEDDED = /*__DATA__*/null;
const PAL = ["#0b3d2e","#15634a","#3f7d6a","#b9975b","#6f9a89","#8aa399","#c9b27a","#a9bdb4","#d8c79b","#2c5446","#e0d3ad","#5b8c7b"];
const B="฿";
const f0=n=>(n<0?"-":"")+B+Math.abs(Math.round(n)).toLocaleString("en-US");
const fM=n=>B+(n/1e6).toFixed(2)+"M";
const pc=n=>(n*100).toFixed(1)+"%";
const sg=n=>n>=0?"+":"";
const md=iso=>{const p=(iso||"").split("-");return p.length===3?(+p[1])+"/"+(+p[2]):iso;};
const MON=["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"];
const dmon=iso=>{const p=(iso||"").split("-");return p.length===3?(+p[2])+" "+(MON[+p[1]-1]||""):iso;};
// "Priced" cell: date of the latest price fetched. Bold if it's today's as-of date
// (freshly updated), amber if the engine flagged it stale (>1 week carried forward).
const priced=(iso,stale,asof)=>{if(!iso)return '<span class="pdate">—</span>';
  const cls=stale?'stale':(iso===asof?'fresh':'');return `<span class="pdate ${cls}" title="${iso}">${md(iso)}</span>`;};
const charts={};

function boot(PAYLOAD){
  const P=PAYLOAD.portfolio, CFG=PAYLOAD.config||{};
  const HIST=(PAYLOAD.history||[]).slice().sort((a,b)=>a.date<b.date?-1:1);
  const FLOWS=PAYLOAD.flows||[];
  const hasHist=HIST.length>=2;
  const H={labels:HIST.map(h=>md(h.date)),value:HIST.map(h=>h.total_value),invested:HIST.map(h=>h.total_invested)};
  const TABS=["Overview","Allocation","Holdings","Performance","Risk","Tax & Lots"];

  if(CFG.accent)document.documentElement.style.setProperty("--accent",CFG.accent);
  if(CFG.accent_soft)document.documentElement.style.setProperty("--accent-2",CFG.accent_soft);
  if(CFG.gold)document.documentElement.style.setProperty("--gold",CFG.gold);

  const firm=CFG.firm_name||"TH Investment";
  document.getElementById("firm").textContent=firm;
  document.getElementById("logo").textContent=(CFG.logo_text||firm.split(/\s+/).map(w=>w[0]).join("")).slice(0,2).toUpperCase();
  document.getElementById("client").textContent=CFG.client_name||"Private Client Portfolio";
  document.getElementById("ref").textContent=CFG.client_ref||"";
  // as-of date + the time the data was last refreshed (from generated_at)
  const _gen=(PAYLOAD.generated_at||"").trim().split(/\s+/);
  const _genTime=_gen.length>=2?_gen.slice(1).join(" "):"";
  document.getElementById("asof").innerHTML=P.as_of+(_genTime?` <span style="color:var(--muted);font-weight:500">${_genTime}</span>`:"");
  document.getElementById("fx").textContent=P.fx_rate;
  document.title=(CFG.report_title||"Private Banking Summary");
  document.getElementById("foot").innerHTML=
    `<div>${CFG.disclaimer||""}</div>
     <div style="margin-top:8px">${CFG.schedule_text||""} · Generated ${PAYLOAD.generated_at||""} ·
       ${P.lot_counts.MF} funds · ${P.lot_counts.Equity} equity · ${P.lot_counts.Crypto} crypto${P.lot_counts.UnitTrust?` · ${P.lot_counts.UnitTrust} unit trust`:''} lots</div>`;

  // pooled holdings
  const pool={};
  P.holdings.forEach(h=>{const k=h.asset_type+"|"+h.name;
    const g=pool[k]||(pool[k]={name:h.name,type:h.asset_type,ac:h.asset_class,geo:h.geography,
      cur:h.currency||"THB",qty:0,price:h.price,inv:0,val:0,asof:"",stale:false});
    g.qty+=h.quantity;g.inv+=h.invested_thb;g.val+=h.value_thb;
    // keep the freshest lot's price date + its staleness for the pooled position
    if(!g.asof||(h.price_asof||"")>g.asof){g.asof=h.price_asof||"";g.stale=!!h.price_stale;}});
  const POOL=Object.values(pool).map(h=>({...h,pnl:h.val-h.inv,pnlpct:h.inv?(h.val-h.inv)/h.inv:0,wt:P.total_value?h.val/P.total_value:0}))
    .sort((a,b)=>b.val-a.val);

  // tabs — top bar (desktop) + bottom nav (mobile), both drive selectTab()
  const nav=document.getElementById("tabnav");nav.innerHTML="";
  const bnav=document.getElementById("bottomnav");bnav.innerHTML="";
  const drawn={};
  const ICONS={
    "Overview":'<rect x="3" y="3" width="7" height="7"/><rect x="14" y="3" width="7" height="7"/><rect x="3" y="14" width="7" height="7"/><rect x="14" y="14" width="7" height="7"/>',
    "Allocation":'<circle cx="12" cy="12" r="9"/><path d="M12 12V3"/><path d="M12 12l7.5 5"/>',
    "Holdings":'<line x1="8" y1="6" x2="21" y2="6"/><line x1="8" y1="12" x2="21" y2="12"/><line x1="8" y1="18" x2="21" y2="18"/><circle cx="3.6" cy="6" r="1"/><circle cx="3.6" cy="12" r="1"/><circle cx="3.6" cy="18" r="1"/>',
    "Performance":'<path d="M3 17l5-6 4 3 6-9"/><path d="M3 21h18"/>',
    "Risk":'<path d="M12 3l8 3v6c0 5-3.6 8-8 9-4.4-1-8-4-8-9V6z"/>',
    "Tax & Lots":'<path d="M6 3h9l4 4v14H6z"/><path d="M9 10h6M9 14h6M9 18h4"/>'
  };
  const BLABEL={"Overview":"Overview","Allocation":"Alloc.","Holdings":"Holdings","Performance":"Perf.","Risk":"Risk","Tax & Lots":"Tax"};
  function selectTab(t){
    document.querySelectorAll(".tab").forEach(x=>x.classList.toggle("active",x.dataset.t===t));
    document.querySelectorAll(".bottomnav button").forEach(x=>x.classList.toggle("active",x.dataset.t===t));
    document.querySelectorAll("section[data-tab]").forEach(s=>s.hidden=s.dataset.tab!==t);
    window.scrollTo(0,0);draw(t);
  }
  TABS.forEach((t,i)=>{
    const b=document.createElement("button");b.className="tab"+(i?"":" active");b.textContent=t;b.dataset.t=t;b.onclick=()=>selectTab(t);nav.appendChild(b);
    const bb=document.createElement("button");bb.className=i?"":"active";bb.dataset.t=t;
    bb.innerHTML='<svg viewBox="0 0 24 24" aria-hidden="true">'+ICONS[t]+'</svg><span>'+BLABEL[t]+'</span>';
    bb.onclick=()=>selectTab(t);bnav.appendChild(bb);
  });
  // tap a mobile holding/lot card to expand its detail
  document.addEventListener("click",e=>{const c=e.target.closest(".lcard");if(c)c.classList.toggle("open");});

  // KPIs
  // true week-over-week: change vs the snapshot ~7 calendar days ago
  function changeOver(days){
    if(HIST.length<2)return null;
    const t=new Date(P.as_of+'T00:00:00'); t.setDate(t.getDate()-days);
    let ref=null; HIST.forEach(h=>{if(new Date(h.date+'T00:00:00')<=t)ref=h;});
    if(!ref)ref=HIST[0];
    return ref.total_value?(P.total_value-ref.total_value)/ref.total_value:null;
  }
  const wow=changeOver(7);
  // day-over-day: the two most recent daily snapshots (markets may not move on
  // weekends, so compare actual adjacent points rather than "exactly yesterday")
  function dayChange(){
    if(HIST.length<2)return null;
    const a=HIST[HIST.length-2], b=HIST[HIST.length-1];
    const mf=(b.mf_value||0)-(a.mf_value||0), eq=(b.eq_value||0)-(a.eq_value||0), cr=(b.crypto_value||0)-(a.crypto_value||0);
    const totAbs=b.total_value-a.total_value;
    // real external cash flows dated within the window (a.date, b.date] — actual
    // money in/out, NOT inferred from cost basis (which drifts daily with FX).
    const win=FLOWS.filter(f=>f.date>a.date&&f.date<=b.date);
    const flow=win.reduce((s,f)=>s+(f.amount||0),0);
    return {prevDate:a.date, valAbs:totAbs, valPct:a.total_value?totAbs/a.total_value:null,
      flow, flowItems:win, marketAbs:totAbs-flow,
      mf, eq, crypto:cr, other:totAbs-mf-eq-cr};
  }
  const dc=dayChange();
  function spark(series){
    if(!series||series.length<2)return"";
    const w=74,h=30,mn=Math.min(...series),mx=Math.max(...series),r=mx-mn||1;
    const pts=series.map((v,i)=>[i/(series.length-1)*w,h-((v-mn)/r)*(h-4)-2]);
    const up=series[series.length-1]>=series[0];
    const d=pts.map((p,i)=>(i?"L":"M")+p[0].toFixed(1)+" "+p[1].toFixed(1)).join(" ");
    return `<svg class="spark" viewBox="0 0 ${w} ${h}"><path d="${d}" fill="none" stroke="${up?'#0f7a44':'#c0392b'}" stroke-width="1.6"/></svg>`;
  }
  const d1=dc?dc.valPct:null;
  const kpis=[
    {label:"Total Portfolio Value",val:f0(P.total_value),d:d1,dl:"1D",spark:H.value},
    {label:"Invested (Cost Basis)",val:f0(P.total_invested),sub:"THB"},
    {label:"Unrealized P&L",val:sg(P.total_pnl)+f0(P.total_pnl),cls:P.total_pnl>=0?"pos":"neg",sub:"since inception"},
    {label:"Total Return",val:sg(P.total_pnl_pct)+pc(P.total_pnl_pct),cls:P.total_pnl_pct>=0?"pos":"neg",d:d1,dl:"1D",spark:H.value},
  ];
  document.getElementById("kpis").innerHTML=kpis.map(k=>`
    <div class="card kpi">
      <div class="label">${k.label}</div>
      <div class="val num ${k.cls||''}">${k.val}</div>
      ${k.d!=null?`<div class="delta ${k.d>=0?'up':'down'}">${k.d>=0?'▲':'▼'} ${pc(Math.abs(k.d))} ${k.dl||'1D'}</div>`
                 :`<div class="mini" style="margin-top:8px">${k.sub||''}</div>`}
      ${k.spark?spark(k.spark):''}
    </div>`).join("");

  // Today's Movement card — headline day-over-day change + asset-bucket breakdown
  (function(){
    const el=document.getElementById("dailyMove");
    if(!dc){el.innerHTML=`<div class="empty" style="padding:16px">Daily movement is building.`
      +`<br><span class="pillnote">Two daily snapshots needed — populates after the next refresh.</span></div>`;return;}
    const up=dc.valAbs>=0;
    const breaks=[["Funds",dc.mf],["Equities",dc.eq],["Crypto",dc.crypto],["Other",dc.other]]
      .map(([n,v])=>Math.abs(v)<1?`<span class="movchip flat">${n} ·฿0</span>`
        :`<span class="movchip ${v>=0?'up':'down'}">${n} ${v>=0?'▲':'▼'} ${f0(Math.abs(v))}</span>`).join("");
    // Only attribute a flow when the Cash Flows sheet has a real dated entry in the
    // window; otherwise the whole move is market & FX (no contribution).
    let note="";
    if(Math.abs(dc.flow)>=1){
      const lbl=(dc.flowItems.find(f=>f.type)||{}).type
        || (dc.flow>=0?"contribution":"withdrawal");
      note=`<div class="mini" style="margin-top:9px">Includes ${sg(dc.flow)}${f0(dc.flow)} ${lbl.toLowerCase()}`
        +` · market &amp; FX ${sg(dc.marketAbs)}${f0(dc.marketAbs)}</div>`;
    }
    el.innerHTML=`<div class="movmain">
        <div>
          <div class="label">Today's Movement</div>
          <div class="movbig num ${up?'pos':'neg'}">${sg(dc.valAbs)}${f0(dc.valAbs)}`
            +`<span class="movpct ${up?'up':'down'}">${up?'▲':'▼'} ${pc(Math.abs(dc.valPct))}</span></div>
          <div class="mini">vs ${dmon(dc.prevDate)}</div>
        </div>
        <div class="movbreak">${breaks}</div>
      </div>${note}`;
  })();

  // allocation bar
  const ac=P.by_asset_class.filter(r=>r.value>0);
  document.getElementById("acTotal").textContent=f0(P.total_value);
  document.getElementById("acBar").innerHTML=ac.map((r,i)=>`<i style="width:${r.pct*100}%;background:${PAL[i%PAL.length]}" title="${r.name} ${pc(r.pct)}"></i>`).join("");
  document.getElementById("acLegend").innerHTML=ac.map((r,i)=>`<div><span class="dot" style="background:${PAL[i%PAL.length]}"></span>${r.name} · <span class="num">${pc(r.pct)}</span></div>`).join("");

  // movers
  const mv=POOL.filter(h=>h.val>50000);
  const up3=[...mv].sort((a,b)=>b.pnlpct-a.pnlpct).slice(0,3);
  const dn3=[...mv].sort((a,b)=>a.pnlpct-b.pnlpct).slice(0,3);
  const mrow=h=>`<div style="display:flex;justify-content:space-between;gap:10px;padding:7px 0;border-bottom:1px solid var(--border)">
    <div style="min-width:0"><span class="name" style="white-space:normal;overflow-wrap:anywhere">${h.name}</span> <span class="sub">${h.type}</span></div>
    <div class="num ${h.pnl>=0?'pos':'neg'}" style="white-space:nowrap;flex-shrink:0">${sg(h.pnlpct)}${pc(h.pnlpct)}</div></div>`;
  document.getElementById("movers").innerHTML=
    `<div class="mini" style="margin:2px 0 4px">Gainers</div>${up3.map(mrow).join("")}
     <div class="mini" style="margin:12px 0 4px">Laggards</div>${dn3.map(mrow).join("")}`;

  // largest holdings
  document.getElementById("ovHold").innerHTML=
    `<thead><tr><th>Holding</th><th>Value</th><th>Wt</th><th>Return</th></tr></thead><tbody>`+
    POOL.slice(0,6).map(h=>`<tr><td><span class="name">${h.name}</span> <span class="sub">${h.type}</span></td>
     <td class="num">${f0(h.val)}</td><td class="num">${pc(h.wt)}</td>
     <td class="num ${h.pnl>=0?'pos':'neg'}">${sg(h.pnlpct)}${pc(h.pnlpct)}</td></tr>`).join("")+`</tbody>`;

  // ---- helpers for tab charts ----
  function donut(id,rows){
    if(charts[id])charts[id].destroy();
    const top=rows.filter(r=>r.value>0);
    charts[id]=new Chart(document.getElementById(id),{type:"doughnut",
      data:{labels:top.map(r=>r.name),datasets:[{data:top.map(r=>r.value),backgroundColor:top.map((_,i)=>PAL[i%PAL.length]),borderColor:"#fff",borderWidth:2}]},
      options:{cutout:"64%",plugins:{legend:{position:"bottom",labels:{boxWidth:9,font:{size:11},padding:9}},
        tooltip:{callbacks:{label:c=>" "+c.label+": "+pc(P.total_value?c.parsed/P.total_value:0)}}}}});
  }
  function bars(elId,rows,max){
    document.getElementById(elId).innerHTML=rows.filter(r=>r.value>0).slice(0,12).map((r,i)=>`
      <div style="display:grid;grid-template-columns:120px 1fr 52px;gap:10px;align-items:center;padding:5px 0">
        <div style="font-size:12.5px">${r.name}</div>
        <div class="wbar"><i style="width:${max?(r.value/max*100).toFixed(0):0}%;background:${PAL[i%PAL.length]}"></i></div>
        <div class="num" style="text-align:right;font-size:12px">${pc(r.pct)}</div></div>`).join("");
  }
  const buildingMsg=`<div class="empty">Performance history is building.<br>
    <span class="pillnote">One weekly snapshot so far — charts populate each Saturday.</span></div>`;

  // holdings table
  let hsort={k:"val",dir:-1},hfilter="All",hq="";
  const seg=document.getElementById("typeSeg");seg.innerHTML="";
  ["All","MF","Equity","Crypto","Unit Trust"].forEach((t,i)=>{const b=document.createElement("button");
    b.className=i?"":"active";b.textContent=t==="MF"?"Funds":t;b.onclick=()=>{hfilter=t;
     [...seg.children].forEach(x=>x.classList.remove("active"));b.classList.add("active");renderHold();};seg.appendChild(b);});
  document.getElementById("search").addEventListener("input",e=>{hq=e.target.value.toLowerCase();renderHold();});
  const HCOLS=[["name","Holding"],["type","Type"],["ac","Class"],["geo","Geo"],["qty","Units"],["price","Price"],
    ["inv","Invested"],["val","Value"],["wt","Wt"],["pnl","P&L"],["pnlpct","P&L %"],["asof","Priced"]];
  function renderHold(){
    let rows=POOL.filter(h=>(hfilter==="All"||h.type===hfilter)&&
      (!hq||(h.name+" "+h.ac+" "+h.geo).toLowerCase().includes(hq)));
    rows.sort((a,b)=>(a[hsort.k]>b[hsort.k]?1:-1)*hsort.dir);
    document.getElementById("holdCount").textContent=rows.length+" of "+POOL.length+" positions";
    const th=HCOLS.map(c=>`<th data-k="${c[0]}">${c[1]}${hsort.k===c[0]?(hsort.dir<0?" ↓":" ↑"):""}</th>`).join("");
    document.getElementById("holdTable").innerHTML=`<thead><tr>${th}</tr></thead><tbody>`+
      rows.map(h=>`<tr>
        <td><span class="name">${h.name}</span></td><td><span class="chip">${h.type}</span></td>
        <td class="sub">${h.ac||""}</td><td class="sub">${h.geo||""}</td>
        <td class="num">${h.qty.toLocaleString("en-US",{maximumFractionDigits:3})}</td>
        <td class="num">${h.price.toLocaleString("en-US",{maximumFractionDigits:2})}</td>
        <td class="num">${f0(h.inv)}</td><td class="num">${f0(h.val)}</td>
        <td class="num">${pc(h.wt)}</td>
        <td class="num ${h.pnl>=0?'pos':'neg'}">${sg(h.pnl)}${f0(h.pnl)}</td>
        <td class="num ${h.pnl>=0?'pos':'neg'}">${sg(h.pnlpct)}${pc(h.pnlpct)}</td>
        <td class="num">${priced(h.asof,h.stale,P.as_of)}</td></tr>`).join("")+`</tbody>`;
    document.querySelectorAll("#holdTable th").forEach(t=>t.onclick=()=>{const k=t.dataset.k;
      if(hsort.k===k)hsort.dir*=-1;else{hsort.k=k;hsort.dir=-1;}renderHold();});
    // mobile cards
    document.getElementById("holdCards").innerHTML=rows.map(h=>{const c=h.pnl>=0?'pos':'neg';return `<div class="lcard">
      <div class="lcard-top"><div class="lcard-nm">${h.name} <span class="chip">${h.type}</span><span class="caret">▾</span></div>
        <div class="lcard-rt"><span class="v num">${f0(h.val)}</span><span class="r ${c}">${sg(h.pnlpct)}${pc(h.pnlpct)}</span></div></div>
      <div class="lcard-more">
        <div><span>Class</span><b>${h.ac||'—'}</b></div><div><span>Geography</span><b>${h.geo||'—'}</b></div>
        <div><span>Units</span><b>${h.qty.toLocaleString("en-US",{maximumFractionDigits:3})}</b></div>
        <div><span>Price</span><b>${h.price.toLocaleString("en-US",{maximumFractionDigits:2})}</b></div>
        <div><span>Invested</span><b>${f0(h.inv)}</b></div><div><span>Weight</span><b>${pc(h.wt)}</b></div>
        <div><span>P&L</span><b class="${c}">${sg(h.pnl)}${f0(h.pnl)}</b></div>
        <div><span>Priced</span><b>${priced(h.asof,h.stale,P.as_of)}</b></div></div></div>`;}).join("");
  }

  // performance
  let perfRange=1825;
  function fpct(x,dec){dec=dec==null?1:dec;return x==null?'—':(x>=0?'+':'')+(x*100).toFixed(dec)+'%';}
  function pcls(x){return x==null?'':x>=0?'pos':'neg';}
  function perfCharts(){
    const PF=PAYLOAD.performance||{}; const pr=PF.period_returns||{}, st=PF.stats||{}, periods=PF.periods||[];
    const pn=pr.portfolio_net||{}, bm=pr.benchmark||{};
    const ytd=pn['YTD'], bytd=bm['YTD'], si=st.annualized_return, sicum=st.since_inception_cum, vol=st.volatility, sh=st.sharpe, dd=st.max_drawdown;
    const days=PF.history_days||0;
    const siVal = si!=null ? fpct(si) : fpct(sicum);
    const siLbl = si!=null ? ((days/365.25).toFixed(1)+' yrs · p.a.') : (sicum!=null?'cumulative · '+days+'d':'building');
    const tiles=[
      ['TWR — YTD',fpct(ytd),(ytd!=null&&bytd!=null)?'vs bmk '+fpct(ytd-bytd):'time-weighted',pcls(ytd)],
      ['Since Inception',siVal,siLbl,pcls(si!=null?si:sicum)],
      ['Money-Weighted (IRR)',fpct(st.irr),st.irr!=null?'incl. cash flows':'needs cash-flow log',pcls(st.irr)],
      ['Volatility',vol==null?'—':(vol*100).toFixed(1)+'%','annualized',''],
      ['Sharpe',sh==null?'—':sh.toFixed(2),sh==null?'needs ≥1yr':'rf '+(CFG.risk_free_pct!=null?CFG.risk_free_pct:2)+'%',''],
      ['Max Drawdown',dd==null?'—':(dd*100).toFixed(1)+'%','since inception',dd==null?'':'neg'],
    ];
    document.getElementById('perfStats').innerHTML=tiles.map(t=>`<div class="card stat"><div class="k">${t[0]}</div><div class="v num ${t[3]}">${t[1]}</div><div class="d">${t[2]}</div></div>`).join('');
    document.getElementById('perfBmkLbl').textContent='Time-weighted · net of fees · benchmark: '+(PF.benchmark_name||'—');
    document.getElementById('perfBmkName').textContent=PF.benchmark_name||'Benchmark';

    const head='<th>Return</th>'+periods.map(p=>`<th>${p}</th>`).join('');
    const rowFor=(obj,lbl,colorRel)=>`<tr><td class="name">${lbl}</td>`+periods.map(p=>{const v=(obj||{})[p];return `<td class="num ${colorRel?pcls(v):''}">${fpct(v)}</td>`}).join('')+'</tr>';
    document.getElementById('perfPeriod').innerHTML=`<thead><tr>${head}</tr></thead><tbody>`+
      rowFor(pn,'Portfolio (net)',false)+rowFor(bm,PF.benchmark_name||'Benchmark',false)+rowFor(pr.relative,'Relative',true)+`</tbody>`;

    const rr=[[si!=null?'Annualized return':'Return (since inception)',fpct(si!=null?si:sicum)],['Volatility',vol==null?'—':(vol*100).toFixed(1)+'%'],['Sharpe ratio',sh==null?'—':sh.toFixed(2)],['Sortino ratio',st.sortino==null?'—':st.sortino.toFixed(2)],['Max drawdown',dd==null?'—':(dd*100).toFixed(1)+'%'],['Best day',fpct(st.best_period)],['Worst day',fpct(st.worst_period)],['% positive days',st.pct_positive==null?'—':Math.round(st.pct_positive*100)+'%']];
    document.getElementById('perfRisk').innerHTML='<tbody>'+rr.map(r=>`<tr><td style="color:var(--ink);font-size:12.5px">${r[0]}</td><td class="num">${r[1]}</td></tr>`).join('')+'</tbody>';

    drawGrowth(); drawDD(); drawCal();
    const sel=document.getElementById('perfRange');
    if(sel&&!sel._wired){sel._wired=1;sel.addEventListener('click',e=>{if(e.target.tagName!=='BUTTON')return;[...sel.children].forEach(b=>b.classList.remove('active'));e.target.classList.add('active');perfRange=+e.target.dataset.r;drawGrowth();});}
    const n=document.getElementById('perfNote');
    if((PF.history_points||0)<25){n.style.display='block';n.innerHTML='Portfolio metrics accrue as daily snapshots build — currently <b>'+(PF.history_points||0)+'</b> day'+((PF.history_points||0)===1?'':'s')+' since inception '+(PF.inception||'')+'. The benchmark and full framework are shown now; the period table, growth line and risk stats fill in over the coming days.';}
    else n.style.display='none';
  }
  function drawGrowth(){
    const PF=PAYLOAD.performance||{}; const g=PF.growth||{portfolio:[],benchmark:[]};
    const asof=new Date((PF.as_of||'')+'T00:00:00'); const cut=new Date(asof); cut.setDate(cut.getDate()-perfRange);
    const inWin=s=>new Date(s.date+'T00:00:00')>=cut;
    let b=(g.benchmark||[]).filter(inWin); if(!b.length)b=g.benchmark||[];
    const bBase=b.length?b[0].level:100; const labels=b.map(s=>s.date);
    const bData=b.map(s=>s.level/bBase*100);
    let pf=(g.portfolio||[]).filter(inWin); const pBase=pf.length?pf[0].level:100;
    const pMap={}; pf.forEach(s=>pMap[s.date]=s.level/pBase*100);
    const pData=labels.map(d=>pMap[d]!=null?pMap[d]:null);
    if(charts.perfGrowth)charts.perfGrowth.destroy();
    charts.perfGrowth=new Chart(document.getElementById('perfGrowth'),{type:'line',data:{labels,datasets:[
      {label:'Portfolio',data:pData,borderColor:'#0b3d2e',backgroundColor:'rgba(11,61,46,.06)',fill:true,borderWidth:2,pointRadius:0,spanGaps:true,tension:.15},
      {label:'Benchmark',data:bData,borderColor:'#b9975b',borderDash:[5,4],fill:false,borderWidth:1.5,pointRadius:0,tension:.15}]},
      options:{plugins:{legend:{display:false}},scales:{y:{grid:{color:'#eef0f2'},ticks:{font:{size:11},color:'#69727b'}},x:{grid:{display:false},ticks:{font:{size:11},color:'#69727b',maxTicksLimit:8,callback:(v,i)=>labels[i]?labels[i].slice(0,7):''}}}}});
  }
  function drawDD(){
    const dd=(PAYLOAD.performance||{}).drawdown||[];
    if(charts.perfDD)charts.perfDD.destroy();
    const wrap=document.getElementById('perfDDWrap');
    if(!dd.length){wrap.innerHTML='<div style="font-size:12px;color:var(--faint);text-align:center;padding-top:46px">builds with daily history</div>';return;}
    if(!document.getElementById('perfDD')){wrap.innerHTML='<canvas id="perfDD"></canvas>';}
    charts.perfDD=new Chart(document.getElementById('perfDD'),{type:'line',data:{labels:dd.map(x=>x.date),datasets:[{data:dd.map(x=>x.dd),borderColor:'#c0392b',backgroundColor:'rgba(192,57,43,.10)',fill:true,borderWidth:1.2,pointRadius:0,tension:.1}]},
      options:{plugins:{legend:{display:false}},scales:{y:{grid:{color:'#eef0f2'},ticks:{font:{size:11},color:'#69727b',callback:v=>v.toFixed(0)+'%'}},x:{grid:{display:false},ticks:{font:{size:11},color:'#69727b',maxTicksLimit:5}}}}});
  }
  function drawCal(){
    const c=(PAYLOAD.performance||{}).calendar||{years:[],portfolio:[],benchmark:[]};
    if(charts.perfCal)charts.perfCal.destroy();
    charts.perfCal=new Chart(document.getElementById('perfCal'),{type:'bar',data:{labels:c.years,datasets:[
      {label:'Portfolio',data:c.portfolio,backgroundColor:'#0b3d2e'},{label:'Benchmark',data:c.benchmark,backgroundColor:'#cdb48a'}]},
      options:{plugins:{legend:{position:'bottom',labels:{boxWidth:10,font:{size:11}}}},scales:{y:{grid:{color:'#eef0f2'},ticks:{font:{size:11},color:'#69727b',callback:v=>(v==null?'':v.toFixed(0)+'%')}},x:{grid:{display:false},ticks:{font:{size:11},color:'#69727b'}}}}});
  }

  // overview mini perf
  function perfMini(){
    if(!hasHist){document.getElementById("ovPerfWrap").innerHTML=buildingMsg;return;}
    if(charts.ovPerf)charts.ovPerf.destroy();
    charts.ovPerf=new Chart(document.getElementById("ovPerf"),{type:"line",
     data:{labels:H.labels,datasets:[
       {label:"Value",data:H.value,borderColor:"#0b3d2e",backgroundColor:"rgba(11,61,46,.07)",fill:true,tension:.3,pointRadius:0,borderWidth:2},
       {label:"Invested",data:H.invested,borderColor:"#b9975b",borderDash:[5,4],fill:false,tension:.3,pointRadius:0,borderWidth:1.5}]},
     options:{plugins:{legend:{position:"bottom",labels:{boxWidth:10,font:{size:11}}}},
       scales:{y:{grid:{color:"#eef0f2"},ticks:{font:{size:11},color:"#69727b",callback:v=>fM(v)}},x:{grid:{display:false},ticks:{font:{size:11},color:"#69727b"}}}}});
  }

  // risk
  function riskViews(){
    const hhi=POOL.reduce((a,h)=>a+h.wt*h.wt,0);
    const effN=hhi?1/hhi:0;
    const top10=POOL.slice(0,10).reduce((a,h)=>a+h.wt,0);
    const top5=POOL.slice(0,5).reduce((a,h)=>a+h.wt,0);
    document.getElementById("riskStats").innerHTML=[
      ["Largest Position",pc(POOL[0]?POOL[0].wt:0),POOL[0]?POOL[0].name:""],
      ["Top 5 Weight",pc(top5),"of portfolio"],
      ["Top 10 Weight",pc(top10),"of portfolio"],
      ["Effective Holdings",effN.toFixed(1),POOL.length+" positions held"],
    ].map(s=>`<div class="card stat"><div class="k">${s[0]}</div><div class="v num">${s[1]}</div><div class="d">${s[2]}</div></div>`).join("");
    document.getElementById("concList").innerHTML=`<ul>`+POOL.slice(0,10).map((h,i)=>`<li>
      <span class="nm">${h.name}</span>
      <span class="wbar"><i style="width:${POOL[0].wt?(h.wt/POOL[0].wt*100).toFixed(0):0}%;background:${PAL[i%PAL.length]}"></i></span>
      <span class="num" style="text-align:right">${pc(h.wt)}</span></li>`).join("")+`</ul>`;
    const fx={};POOL.forEach(h=>{const c=h.cur==="THB"?"THB":h.cur==="SGD"?"SGD":"USD / USDT";fx[c]=(fx[c]||0)+h.val;});
    const fxRows=Object.entries(fx).sort((a,b)=>b[1]-a[1]);
    const fxcol={"THB":"#0b3d2e","USD / USDT":"#b9975b","SGD":"#3f7d6a"};
    document.getElementById("fxBar").innerHTML=fxRows.map(([k,v])=>`<i style="width:${P.total_value?v/P.total_value*100:0}%;background:${fxcol[k]||'#6f9a89'}"></i>`).join("");
    document.getElementById("fxLegend").innerHTML=fxRows.map(([k,v])=>`<div><span class="dot" style="background:${fxcol[k]||'#6f9a89'}"></span>${k} · <span class="num">${pc(P.total_value?v/P.total_value:0)}</span> · ${f0(v)}</div>`).join("");
    const band={"Equity":"Growth","Digital Assets":"Growth","Multi-Asset":"Balanced","Alternatives":"Balanced","Fixed Income":"Defensive","Cash & Equivalents":"Cash"};
    const bcol={Growth:"#0b3d2e",Balanced:"#3f7d6a",Defensive:"#b9975b",Cash:"#aeb6bd"};
    const agg={};P.by_asset_class.forEach(r=>{const b=band[r.name]||"Balanced";agg[b]=(agg[b]||0)+r.value;});
    const order=["Growth","Balanced","Defensive","Cash"].filter(b=>agg[b]);
    document.getElementById("riskRibbon").innerHTML=order.map(b=>`<i style="width:${P.total_value?agg[b]/P.total_value*100:0}%;background:${bcol[b]}">${(P.total_value?agg[b]/P.total_value*100:0).toFixed(0)}%</i>`).join("");
    document.getElementById("riskLegend").innerHTML=order.map(b=>`<div><span class="dot" style="background:${bcol[b]}"></span>${b} · <span class="num">${pc(P.total_value?agg[b]/P.total_value:0)}</span></div>`).join("");
  }

  function draw(tab){
    if(tab==="Overview"&&!drawn.ov){perfMini();drawn.ov=1;}
    if(tab==="Allocation"&&!drawn.al){donut("acDonut",P.by_asset_class);donut("taxDonut",P.by_tax_status);
      bars("geoBars",P.by_geography,Math.max(...P.by_geography.map(r=>r.value),1));
      bars("themeBars",P.by_theme,Math.max(...P.by_theme.map(r=>r.value),1));
      document.getElementById("acTable").innerHTML=`<thead><tr><th>Asset Class</th><th>Invested</th><th>Value</th><th>Wt</th><th>P&L</th></tr></thead><tbody>`+
        P.by_asset_class.filter(r=>r.value>0).map(r=>`<tr><td class="name">${r.name}</td><td class="num">${f0(r.invested)}</td>
        <td class="num">${f0(r.value)}</td><td class="num">${pc(r.pct)}</td>
        <td class="num ${r.pnl>=0?'pos':'neg'}">${sg(r.pnl)}${f0(r.pnl)}</td></tr>`).join("")+`</tbody>`;drawn.al=1;}
    if(tab==="Holdings"&&!drawn.ho){renderHold();drawn.ho=1;}
    if(tab==="Performance"&&!drawn.pe){perfCharts();drawn.pe=1;}
    if(tab==="Risk"&&!drawn.ri){riskViews();drawn.ri=1;}
    if(tab==="Tax & Lots"&&!drawn.tx){renderTax();drawn.tx=1;}
  }

  function renderTax(){
    const T=PAYLOAD.tax||{}; if(T.error||!T.totals){document.getElementById('taxNote').textContent='Tax view unavailable.';return;}
    const tt=T.totals;
    const tiles=[
      ['Unrealized Gain/Loss',sg(tt.unrealized)+f0(tt.unrealized),sg(tt.unrealized_pct)+pc(tt.unrealized_pct)+' on cost',tt.unrealized>=0?'pos':'neg'],
      ['Realized YTD',tt.realized_count?sg(tt.realized_ytd)+f0(tt.realized_ytd):'฿0',tt.realized_count?tt.realized_count+' disposal(s)':'no disposals logged',tt.realized_ytd>=0?'':'neg'],
      ['Tax-Advantaged',f0(tt.tax_adv_value),pc(tt.tax_adv_pct)+' of portfolio',''],
      ['Locked (in wrappers)',f0(tt.locked_value),'until maturity','neg'],
      ['Available Now',f0(tt.available_value),'liquid / matured','pos'],
      ['Cost Basis',f0(tt.cost),'total invested',''],
    ];
    document.getElementById('taxStats').innerHTML=tiles.map(t=>`<div class="card stat"><div class="k">${t[0]}</div><div class="v num ${t[3]}">${t[1]}</div><div class="d">${t[2]}</div></div>`).join('');

    // maturity bar
    const mat=T.maturity||[];
    if(charts.taxMaturity)charts.taxMaturity.destroy();
    if(mat.length){
      const fAmt=v=>v>=1e6?'฿'+(v/1e6).toFixed(2)+'M':'฿'+Math.round(v/1e3)+'k';
      const barLabels={id:'barLabels',afterDatasetsDraw(ch){const ctx=ch.ctx,meta=ch.getDatasetMeta(0);
        ctx.save();ctx.font='600 11.5px -apple-system,BlinkMacSystemFont,Segoe UI,Roboto,sans-serif';ctx.fillStyle='#14181c';ctx.textAlign='center';ctx.textBaseline='bottom';
        meta.data.forEach((bar,i)=>{const v=ch.data.datasets[0].data[i];if(v!=null)ctx.fillText(fAmt(v),bar.x,bar.y-5);});ctx.restore();}};
      charts.taxMaturity=new Chart(document.getElementById('taxMaturity'),{type:'bar',
        data:{labels:mat.map(m=>m.year),datasets:[{data:mat.map(m=>m.value),backgroundColor:'#15634a',maxBarThickness:80,borderRadius:3}]},
        options:{layout:{padding:{top:18}},plugins:{legend:{display:false},tooltip:{callbacks:{label:c=>' '+f0(c.parsed)}}},
          scales:{y:{grace:'12%',grid:{color:'#eef0f2'},ticks:{font:{size:11},color:'#69727b',maxTicksLimit:5,callback:v=>'฿'+(v/1e6).toFixed(1)+'M'}},x:{grid:{display:false},ticks:{font:{size:12},color:'#69727b'}}}},
        plugins:[barLabels]});
    } else { document.getElementById('taxMaturity').parentElement.innerHTML='<div style="text-align:center;color:var(--faint);font-size:12px;padding-top:60px">No locked wrapper holdings</div>'; }
    document.getElementById('taxAvailNote').textContent='Available now in wrappers (matured, e.g. LTF): '+f0(tt.available_in_wrapper)+'.';

    // by wrapper table
    document.getElementById('taxWrapper').innerHTML='<thead><tr><th>Wrapper</th><th>Value</th><th>Wt</th><th>Unrealized</th><th>Locked</th><th>Next</th></tr></thead><tbody>'+
      (T.by_wrapper||[]).map(w=>`<tr><td class="name">${w.wrapper}${w.advantaged?' <span class="chip">tax</span>':''}</td>
        <td class="num">${f0(w.value)}</td><td class="num">${pc(w.pct)}</td>
        <td class="num ${w.unrealized>=0?'pos':'neg'}">${sg(w.unrealized)}${f0(w.unrealized)}</td>
        <td class="num">${w.locked?f0(w.locked):'—'}</td><td class="num">${w.next_unlock||'—'}</td></tr>`).join('')+'</tbody>';

    // lots table
    let tq='',tf='All';
    const seg=document.getElementById('taxSeg');
    const wrappers=['All',...[...new Set((T.lots||[]).map(l=>l.wrapper))]];
    seg.innerHTML=wrappers.map((w,i)=>`<button class="${i?'':'active'}" data-w="${w}">${w}</button>`).join('');
    function renderLots(){
      let rows=(T.lots||[]).filter(l=>(tf==='All'||l.wrapper===tf)&&(!tq||(l.name+' '+l.wrapper+' '+l.type).toLowerCase().includes(tq)));
      document.getElementById('taxCount').textContent=rows.length+' of '+(T.lots||[]).length+' lots';
      document.getElementById('taxLots').innerHTML='<thead><tr><th>Holding</th><th>Wrapper</th><th>Acq.</th><th>Held</th><th>Cost</th><th>Value</th><th>Unrealized</th><th>%</th><th>Sellable</th><th>Status</th><th>Priced</th></tr></thead><tbody>'+
        rows.map(l=>`<tr><td><span class="name">${l.name}</span> <span class="sub">${l.type}</span></td>
          <td class="sub">${l.wrapper}</td><td class="sub">${l.acq_date}</td><td class="num sub">${(l.holding_days/30.44).toFixed(1)}mo</td>
          <td class="num">${f0(l.cost)}</td><td class="num">${f0(l.value)}</td>
          <td class="num ${l.unrealized>=0?'pos':'neg'}">${sg(l.unrealized)}${f0(l.unrealized)}</td>
          <td class="num ${l.unrealized>=0?'pos':'neg'}">${sg(l.unrealized_pct)}${pc(l.unrealized_pct)}</td>
          <td class="num sub">${l.sellable_year||'—'}</td>
          <td><span class="chip" style="${l.status==='Locked'?'color:var(--neg);border-color:#f0cfcb;background:var(--neg-bg)':'color:var(--pos);border-color:#bfe0cd;background:var(--pos-bg)'}">${l.status}</span></td>
          <td class="num">${priced(l.price_asof,l.price_stale,T.as_of)}</td></tr>`).join('')+'</tbody>';
      // mobile cards
      document.getElementById('taxCards').innerHTML=rows.map(l=>{const c=l.unrealized>=0?'pos':'neg';
        const stc=l.status==='Locked'?'color:var(--neg);border-color:#f0cfcb;background:var(--neg-bg)':'color:var(--pos);border-color:#bfe0cd;background:var(--pos-bg)';
        return `<div class="lcard"><div class="lcard-top">
          <div class="lcard-nm">${l.name} <span class="chip" style="${stc}">${l.status}</span><span class="caret">▾</span></div>
          <div class="lcard-rt"><span class="v num">${f0(l.value)}</span><span class="r ${c}">${sg(l.unrealized_pct)}${pc(l.unrealized_pct)}</span></div></div>
        <div class="lcard-more">
          <div><span>Wrapper</span><b>${l.wrapper}</b></div><div><span>Type</span><b>${l.type}</b></div>
          <div><span>Acquired</span><b>${l.acq_date}</b></div><div><span>Held</span><b>${(l.holding_days/30.44).toFixed(1)}mo</b></div>
          <div><span>Cost</span><b>${f0(l.cost)}</b></div><div><span>Unrealized</span><b class="${c}">${sg(l.unrealized)}${f0(l.unrealized)}</b></div>
          <div><span>Sellable</span><b>${l.sellable_year||'—'}</b></div>
          <div><span>Priced</span><b>${priced(l.price_asof,l.price_stale,T.as_of)}</b></div></div></div>`;}).join('');
    }
    seg.onclick=e=>{if(e.target.tagName!=='BUTTON')return;[...seg.children].forEach(b=>b.classList.remove('active'));e.target.classList.add('active');tf=e.target.dataset.w;renderLots();};
    document.getElementById('taxSearch').oninput=e=>{tq=e.target.value.toLowerCase();renderLots();};
    renderLots();
    document.getElementById('taxNote').innerHTML='Holding periods run from '+T.inception+' (assumed acquisition for all lots). Lock-up maturity uses each fund’s actual Sellable Year. “Priced” is the date of the latest price/NAV fetched (bold = today, amber = over a week old). Realized gains populate from the Realized sheet. Not tax advice.';
  }
  draw("Overview");
}

fetch('data.json',{cache:'no-store'}).then(r=>r.ok?r.json():Promise.reject()).then(boot).catch(()=>{ if(EMBEDDED) boot(EMBEDDED); });

// PWA: register the offline service worker (https / localhost only)
if('serviceWorker' in navigator){window.addEventListener('load',()=>{navigator.serviceWorker.register('sw.js').catch(()=>{});});}
</script>
</body>
</html>
"""


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", default=str(WORKBOOK_DEFAULT))
    args = ap.parse_args()
    out = build(args.workbook)
    pf = out["portfolio"]
    print(f"Built docs/index.html + data.json (v2) — total ฿{pf['total_value']:,.0f}, "
          f"{len(out['history'])} history point(s).")
