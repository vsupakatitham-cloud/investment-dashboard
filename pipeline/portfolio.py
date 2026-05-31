"""
portfolio.py — Core valuation engine for the TH Investment private-banking workbook.

Reads the lot-level Excel workbook and reproduces every Dashboard formula in
Python so the figures can be rendered to an online dashboard without Excel.

Data model (mirrors the workbook):
  * FX & Assumptions : As-of date + USD/THB spot (single source of truth for FX).
  * Reference        : key "Type|Name|Custodian" -> asset class / sub-class /
                       geography / theme / tax status / currency.
  * Prices           : key "Type|Name|Custodian" -> current price/NAV + currency.
  * MF / Equities / Crypto "- Lots" : one row per purchase lot (raw inputs).

THB convention: THB holdings multiply by 1; USD/USDT holdings multiply by the
single FX rate cell — exactly as the workbook does.
"""
from __future__ import annotations

import datetime as _dt
import re
from dataclasses import dataclass, field, asdict
from pathlib import Path

import openpyxl

WORKBOOK_DEFAULT = Path(__file__).resolve().parent.parent / "TH Investment - Private Banking Summary.xlsx"

# Data starts at row 5 in every Lots sheet; the table ends at a "TOTAL" row.
LOT_FIRST = 5


def _data_rows(ws, name_col=3, first=LOT_FIRST):
    """Row indices holding real lots — scans from `first` to the TOTAL row.

    Stops at the row whose column A is "TOTAL" (or after a run of blank names),
    so lots can be added freely above the TOTAL row without touching any range.
    """
    rows, blanks, r = [], 0, first
    while r <= ws.max_row:
        a = ws.cell(r, 1).value
        if isinstance(a, str) and a.strip().upper() == "TOTAL":
            break
        if ws.cell(r, name_col).value in (None, ""):
            blanks += 1
            if blanks >= 3:
                break
        else:
            blanks = 0
            rows.append(r)
        r += 1
    return rows


_ARITH_RE = re.compile(r"^[0-9.+\-*/() ]+$")


def _num(v, default=0.0):
    """Coerce a cell to a number. Handles plain numbers and simple arithmetic
    typed as a formula in an input cell (e.g. "=254.87+2584.77"), which openpyxl
    returns as a raw string when reading formulas."""
    if v is None or v == "":
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s.startswith("="):
        s = s[1:].strip()
    try:
        return float(s)
    except ValueError:
        if _ARITH_RE.match(s):                       # safe: digits/operators only
            try:
                return float(eval(s, {"__builtins__": {}}, {}))
            except Exception:
                return default
        return default


@dataclass
class Holding:
    asset_type: str          # MF | Equity | Crypto
    name: str                # fund / ticker / coin
    custodian: str           # AMC / broker / exchange
    asset_class: str = ""
    sub_class: str = ""
    geography: str = ""
    theme: str = ""
    tax_status: str = ""
    currency: str = "THB"
    quantity: float = 0.0
    avg_cost: float = 0.0
    price: float = 0.0
    fx: float = 1.0
    invested_thb: float = 0.0
    value_thb: float = 0.0
    pnl_thb: float = 0.0
    pnl_pct: float = 0.0
    lots: int = 1
    price_asof: str = ""
    price_stale: bool = False


@dataclass
class Portfolio:
    as_of: str
    fx_rate: float
    reporting_ccy: str
    total_value: float
    total_invested: float
    total_pnl: float
    total_pnl_pct: float
    mf_value: float
    eq_value: float
    crypto_value: float
    lot_counts: dict
    holdings: list = field(default_factory=list)
    by_asset_class: list = field(default_factory=list)
    by_geography: list = field(default_factory=list)
    by_tax_status: list = field(default_factory=list)
    by_theme: list = field(default_factory=list)
    top_holdings: list = field(default_factory=list)
    stale_prices: list = field(default_factory=list)

    def to_dict(self):
        d = asdict(self)
        return d


def _key(asset_type, name, custodian):
    return f"{asset_type}|{name}|{custodian}"


def load_portfolio(path: Path | str = WORKBOOK_DEFAULT) -> Portfolio:
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=False)

    # --- FX & Assumptions -------------------------------------------------
    fx_ws = wb["FX & Assumptions"]
    as_of_raw = fx_ws["C5"].value
    if isinstance(as_of_raw, _dt.datetime):
        as_of = as_of_raw.date().isoformat()
    else:
        as_of = str(as_of_raw)
    fx_rate = _num(fx_ws["C6"].value, 0.0)
    sgd_rate = _num(fx_ws["C9"].value, 0.0)        # SGD/THB (added for unit trusts)
    reporting_ccy = fx_ws["C7"].value or "THB"

    def fxmult(ccy):
        c = (ccy or "THB").upper()
        if c == "THB":
            return 1.0
        if c == "SGD":
            return sgd_rate
        return fx_rate                              # USD / USDT

    # --- Reference map ----------------------------------------------------
    ref = {}
    rws = wb["Reference"]
    for r in range(5, rws.max_row + 1):
        typ = rws.cell(r, 2).value
        nm = rws.cell(r, 3).value
        cust = rws.cell(r, 4).value
        if not typ or not nm:
            continue
        ref[f"{typ}|{nm}|{cust}"] = dict(
            asset_class=rws.cell(r, 5).value or "",
            sub_class=rws.cell(r, 6).value or "",
            geography=rws.cell(r, 7).value or "",
            theme=rws.cell(r, 8).value or "",
            tax_status=rws.cell(r, 9).value or "",
            currency=rws.cell(r, 10).value or "THB",
        )

    # --- Prices map -------------------------------------------------------
    prices = {}
    pws = wb["Prices"]
    for r in range(5, pws.max_row + 1):
        typ = pws.cell(r, 2).value
        nm = pws.cell(r, 3).value
        cust = pws.cell(r, 4).value
        if not typ or not nm:
            continue
        last = pws.cell(r, 7).value
        if isinstance(last, _dt.datetime):
            last = last.date().isoformat()
        prices[f"{typ}|{nm}|{cust}"] = dict(
            price=_num(pws.cell(r, 5).value, 0.0),
            currency=pws.cell(r, 6).value or "THB",
            last_update=str(last) if last else "",
        )

    holdings: list[Holding] = []

    def lookup_price(typ, nm, cust):
        p = prices.get(f"{typ}|{nm}|{cust}")
        if not p:
            return 0.0, "", True
        stale = p["price"] == 0.0
        if p["last_update"] and as_of:
            try:
                days = (_dt.date.fromisoformat(as_of) - _dt.date.fromisoformat(p["last_update"])).days
                stale = stale or days > 8  # only flag prices over a week out of date
            except ValueError:
                pass
        return p["price"], p["last_update"], stale

    def lookup_ref(typ, nm, cust):
        return ref.get(f"{typ}|{nm}|{cust}", {})

    # --- Mutual Fund lots -------------------------------------------------
    mws = wb["MF - Lots"]
    for r in _data_rows(mws, name_col=3):
        amc = mws.cell(r, 2).value
        fund = mws.cell(r, 3).value
        if not fund:
            continue
        units = _num(mws.cell(r, 12).value)        # L
        avg_cost = _num(mws.cell(r, 13).value)     # M
        invested = _num(mws.cell(r, 15).value)     # O (stored)
        nav, asof, stale = lookup_price("Mutual Fund", fund, amc)
        meta = lookup_ref("Mutual Fund", fund, amc)
        value = units * nav
        holdings.append(Holding(
            asset_type="MF", name=fund, custodian=amc or "",
            asset_class=meta.get("asset_class", ""), sub_class=meta.get("sub_class", ""),
            geography=meta.get("geography", ""), theme=meta.get("theme", ""),
            tax_status=meta.get("tax_status", ""), currency=meta.get("currency", "THB"),
            quantity=units, avg_cost=avg_cost, price=nav, fx=1.0,
            invested_thb=invested, value_thb=value, pnl_thb=value - invested,
            pnl_pct=((value - invested) / invested) if invested else 0.0,
            price_asof=asof, price_stale=stale,
        ))

    # --- Equity lots ------------------------------------------------------
    ews = wb["Equities - Lots"]
    for r in _data_rows(ews, name_col=3):
        broker = ews.cell(r, 2).value
        ticker = ews.cell(r, 3).value
        if not ticker:
            continue
        ccy = ews.cell(r, 4).value or "THB"
        shares = _num(ews.cell(r, 8).value)        # H
        avg_cost = _num(ews.cell(r, 9).value)      # I (native)
        price, asof, stale = lookup_price("Equity", ticker, broker)
        # Cash is held as Shares=balance, AvgCost=1 (no market price) — value it
        # at its balance so it isn't dropped to zero or flagged stale.
        if str(ticker).strip().lower() == "cash":
            price, stale, asof = (avg_cost or 1.0), False, as_of
        meta = lookup_ref("Equity", ticker, broker)
        fx = fxmult(ccy)
        invested = shares * avg_cost * fx
        value = shares * price * fx
        holdings.append(Holding(
            asset_type="Equity", name=ticker, custodian=broker or "",
            asset_class=meta.get("asset_class", "Equity"), sub_class=meta.get("sub_class", ""),
            geography=meta.get("geography", ""), theme=meta.get("theme", ""),
            tax_status=meta.get("tax_status", "Taxable"), currency=ccy,
            quantity=shares, avg_cost=avg_cost, price=price, fx=fx,
            invested_thb=invested, value_thb=value, pnl_thb=value - invested,
            pnl_pct=((value - invested) / invested) if invested else 0.0,
            price_asof=asof, price_stale=stale,
        ))

    # --- Crypto lots ------------------------------------------------------
    cws = wb["Crypto - Lots"]
    for r in _data_rows(cws, name_col=3):
        exch = cws.cell(r, 2).value
        coin = cws.cell(r, 3).value
        if not coin:
            continue
        ccy = cws.cell(r, 4).value or "USD"
        qty = _num(cws.cell(r, 6).value)           # F
        avg_cost = _num(cws.cell(r, 7).value)      # G
        price, asof, stale = lookup_price("Crypto", coin, exch)
        meta = lookup_ref("Crypto", coin, exch)
        fx = fxmult(ccy)
        invested = qty * avg_cost * fx
        value = qty * price * fx
        holdings.append(Holding(
            asset_type="Crypto", name=coin, custodian=exch or "",
            asset_class=meta.get("asset_class", "Digital Assets"),
            sub_class=meta.get("sub_class", ""), geography=meta.get("geography", "Global"),
            theme=meta.get("theme", "Crypto (All)"), tax_status=meta.get("tax_status", "Taxable"),
            currency=ccy, quantity=qty, avg_cost=avg_cost, price=price, fx=fx,
            invested_thb=invested, value_thb=value, pnl_thb=value - invested,
            pnl_pct=((value - invested) / invested) if invested else 0.0,
            price_asof=asof, price_stale=stale,
        ))

    # --- SGD Unit Trusts (self-contained input sheet) ---------------------
    if "Unit Trust (SGD)" in wb.sheetnames:
        uws = wb["Unit Trust (SGD)"]
        for r in _data_rows(uws, name_col=3):       # Fund Name in column C
            platform = uws.cell(r, 2).value
            fund = uws.cell(r, 3).value
            if not fund:
                continue
            units = _num(uws.cell(r, 8).value)      # H
            avg_cost = _num(uws.cell(r, 9).value)   # I (SGD)
            nav = _num(uws.cell(r, 10).value)       # J (SGD)
            fxm = sgd_rate
            invested = units * avg_cost * fxm
            value = units * nav * fxm
            holdings.append(Holding(
                asset_type="Unit Trust", name=fund, custodian=platform or "",
                asset_class=uws.cell(r, 4).value or "", sub_class="",
                geography=uws.cell(r, 6).value or "", theme=uws.cell(r, 5).value or "",
                tax_status=uws.cell(r, 7).value or "Taxable", currency="SGD",
                quantity=units, avg_cost=avg_cost, price=nav, fx=fxm,
                invested_thb=invested, value_thb=value, pnl_thb=value - invested,
                pnl_pct=((value - invested) / invested) if invested else 0.0,
                price_asof=as_of, price_stale=(nav == 0 and units > 0),
            ))

    return _aggregate(holdings, as_of, fx_rate, reporting_ccy)


def _group(holdings, keyfn):
    out = {}
    for h in holdings:
        k = keyfn(h) or "Unclassified"
        g = out.setdefault(k, dict(name=k, invested=0.0, value=0.0, pnl=0.0))
        g["invested"] += h.invested_thb
        g["value"] += h.value_thb
        g["pnl"] += h.pnl_thb
    return out


def _aggregate(holdings, as_of, fx_rate, reporting_ccy) -> Portfolio:
    total_value = sum(h.value_thb for h in holdings)
    total_invested = sum(h.invested_thb for h in holdings)
    total_pnl = total_value - total_invested

    mf_value = sum(h.value_thb for h in holdings if h.asset_type == "MF")
    eq_value = sum(h.value_thb for h in holdings if h.asset_type == "Equity")
    cr_value = sum(h.value_thb for h in holdings if h.asset_type == "Crypto")

    def finalize(groups, denom):
        rows = []
        for g in groups.values():
            g = dict(g)
            g["pct"] = (g["value"] / denom) if denom else 0.0
            g["pnl_pct"] = (g["pnl"] / g["invested"]) if g["invested"] else 0.0
            rows.append(g)
        rows.sort(key=lambda x: x["value"], reverse=True)
        return rows

    by_ac = finalize(_group(holdings, lambda h: h.asset_class), total_value)
    by_geo = finalize(_group(holdings, lambda h: h.geography), total_value)
    by_tax = finalize(_group(holdings, lambda h: h.tax_status), total_value)
    by_theme = finalize(_group(holdings, lambda h: h.theme), total_value)

    # Pooled holdings (aggregate lots by name+custodian) for the top-10 table.
    pooled = {}
    for h in holdings:
        k = (h.asset_type, h.name, h.custodian)
        p = pooled.setdefault(k, dict(
            asset_type=h.asset_type, name=h.name, custodian=h.custodian,
            asset_class=h.asset_class, geography=h.geography, theme=h.theme,
            invested=0.0, value=0.0, pnl=0.0, lots=0))
        p["invested"] += h.invested_thb
        p["value"] += h.value_thb
        p["pnl"] += h.pnl_thb
        p["lots"] += 1
    pooled_rows = list(pooled.values())
    for p in pooled_rows:
        p["pct"] = (p["value"] / total_value) if total_value else 0.0
        p["pnl_pct"] = (p["pnl"] / p["invested"]) if p["invested"] else 0.0
    pooled_rows.sort(key=lambda x: x["value"], reverse=True)
    top10 = pooled_rows[:10]

    stale = sorted({f"{h.asset_type}|{h.name}|{h.custodian}"
                    for h in holdings if h.price_stale and h.value_thb > 0})

    return Portfolio(
        as_of=as_of, fx_rate=fx_rate, reporting_ccy=reporting_ccy,
        total_value=total_value, total_invested=total_invested,
        total_pnl=total_pnl,
        total_pnl_pct=(total_pnl / total_invested) if total_invested else 0.0,
        mf_value=mf_value, eq_value=eq_value, crypto_value=cr_value,
        lot_counts=dict(
            MF=sum(1 for h in holdings if h.asset_type == "MF"),
            Equity=sum(1 for h in holdings if h.asset_type == "Equity"),
            Crypto=sum(1 for h in holdings if h.asset_type == "Crypto"),
            UnitTrust=sum(1 for h in holdings if h.asset_type == "Unit Trust"),
        ),
        holdings=[asdict(h) for h in holdings],
        by_asset_class=by_ac, by_geography=by_geo, by_tax_status=by_tax,
        by_theme=by_theme, top_holdings=top10, stale_prices=stale,
    )


if __name__ == "__main__":
    p = load_portfolio()
    print(f"As of           : {p.as_of}")
    print(f"FX USD/THB      : {p.fx_rate}")
    print(f"Lots MF/EQ/CR   : {p.lot_counts}")
    print(f"Total value THB : {p.total_value:,.0f}")
    print(f"Total invested  : {p.total_invested:,.0f}")
    print(f"Unrealized P&L  : {p.total_pnl:,.0f}  ({p.total_pnl_pct:.2%})")
    print(f"  MF / EQ / CR  : {p.mf_value:,.0f} / {p.eq_value:,.0f} / {p.crypto_value:,.0f}")
    print(f"Stale prices    : {len(p.stale_prices)}")
    print("Top 5 holdings  :")
    for h in p.top_holdings[:5]:
        print(f"   {h['name']:<14} {h['value']:>14,.0f}  {h['pct']:.1%}")
