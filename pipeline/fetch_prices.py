"""
fetch_prices.py — Auto-fetch live prices into the workbook's Prices sheet.

Sources:
  * USD/THB FX      : Yahoo Finance (THB=X)
  * Equities        : Yahoo Finance (Thai tickers -> ".BK", US tickers as-is)
  * Crypto          : CoinGecko simple-price API (USD), THB rows = USD x FX
  * Thai mutual fund NAVs : NOT auto-fetchable here -> value is carried forward
                            and the row is flagged for manual update.

Only the price column (E), the "Last Update" column (G), and the FX / As-of
cells are written; every formula in the workbook is preserved. Network calls
are best-effort: any symbol that fails keeps its previous price and is reported.

Usage:
    python3 pipeline/fetch_prices.py [--asof YYYY-MM-DD] [--workbook PATH] [--offline]
"""
from __future__ import annotations

import argparse
import datetime as _dt
import json
import re
import sys
import time
import urllib.request
import urllib.error
from pathlib import Path

import openpyxl

import fetch_thai_nav

WORKBOOK_DEFAULT = Path(__file__).resolve().parent.parent / "TH Investment - Private Banking Summary.xlsx"
UA = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15) AppleWebKit/537.36"}

# CoinGecko id map for the coins held. Symbols not listed fall back to carry-forward.
COINGECKO_IDS = {
    "BTC": "bitcoin", "ETH": "ethereum", "SOL": "solana", "BNB": "binancecoin",
    "AAVE": "aave", "ARB": "arbitrum", "SUI": "sui", "DOGE": "dogecoin",
    "COMP": "compound-governance-token", "USDT": "tether",
    "PENGU": "pudgy-penguins", "PLUME": "plume", "TRUMP": "official-trump",
    "VELO": "velo", "VELODROME": "velodrome-finance", "KUB": "bitkub-coin",
}
# Optional user overrides/additions (written by add_holding.py): {SYMBOL: coingecko_id}
_OVERRIDE = Path(__file__).resolve().parent / "coingecko_ids.json"
if _OVERRIDE.exists():
    try:
        COINGECKO_IDS.update({k.upper(): v for k, v in json.loads(_OVERRIDE.read_text()).items()})
    except Exception:
        pass
STABLECOINS = {"USDT": 1.0}


def _get_json(url, timeout=15):
    req = urllib.request.Request(url, headers=UA)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return json.loads(r.read().decode())


def yahoo_quote(symbol, timeout=15):
    """(price, prev_close) for a Yahoo symbol — prev_close is the prior trading
    day's close, for the daily-move calc. Either may be None. The 5d/1d series we
    already pull carries both, so this is free."""
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?range=5d&interval=1d"
    try:
        data = _get_json(url, timeout)
        res = data["chart"]["result"][0]
        meta = res["meta"]
        price = meta.get("regularMarketPrice") or meta.get("previousClose")
        price = float(price) if price is not None else None
        prev = None
        # prefer the second-to-last daily close vs the current price; fall back to
        # meta.chartPreviousClose / previousClose when the series is thin.
        try:
            closes = [c for c in res["indicators"]["quote"][0]["close"] if c is not None]
            if price is not None and closes:
                # last close may equal today's price; the prior distinct close is "prev"
                prior = [c for c in closes if abs(c - price) > 1e-9] or closes
                prev = float(prior[-1]) if len(closes) >= 2 else None
        except Exception:
            prev = None
        if prev is None:
            pc = meta.get("chartPreviousClose") or meta.get("previousClose")
            prev = float(pc) if pc is not None else None
        return price, prev
    except Exception:
        return None, None


def yahoo_price(symbol, timeout=15):
    """Latest regular-market price for a Yahoo symbol, or None (FX/simple uses)."""
    return yahoo_quote(symbol, timeout)[0]


def yahoo_symbol_for_isin(isin, timeout=15):
    """Resolve a 12-char ISIN to a Yahoo quote symbol via the search API."""
    url = f"https://query2.finance.yahoo.com/v1/finance/search?q={isin}&quotesCount=4&newsCount=0"
    try:
        data = _get_json(url, timeout)
        for q in data.get("quotes", []):
            if q.get("symbol"):
                return q["symbol"]
    except Exception:
        pass
    return None


def coingecko_prices(ids, timeout=20):
    """{id: {"usd": price, "prev": usd_24h_ago}} — prev derived from the free
    24h-change field (price / (1 + chg/100))."""
    if not ids:
        return {}
    url = ("https://api.coingecko.com/api/v3/simple/price?ids=" + ",".join(sorted(set(ids)))
           + "&vs_currencies=usd&include_24hr_change=true")
    try:
        data = _get_json(url, timeout)
        out = {}
        for k, v in data.items():
            if "usd" not in v:
                continue
            usd = float(v["usd"])
            chg = v.get("usd_24h_change")
            prev = usd / (1 + float(chg) / 100.0) if chg not in (None, "") and (1 + float(chg) / 100.0) != 0 else None
            out[k] = {"usd": usd, "prev": prev}
        return out
    except Exception:
        return {}


def yahoo_symbol(ticker, currency):
    t = (ticker or "").strip().upper()
    if (currency or "THB").upper() == "THB":
        return f"{t.replace('-', '-')}.BK"   # Thai-listed equity/warrant
    return t                                  # US-listed


def fetch_all(workbook=WORKBOOK_DEFAULT, asof=None, offline=False):
    workbook = Path(workbook)
    wb = openpyxl.load_workbook(workbook, data_only=False)
    asof_date = _dt.date.fromisoformat(asof) if asof else None

    report = {"fx": None, "updated": [], "carried": [], "failed": [], "sec": ""}

    # ---- FX --------------------------------------------------------------
    fx_ws = wb["FX & Assumptions"]
    fx_old = float(fx_ws["C6"].value or 0)
    fx_rate = None
    if not offline:
        fx_rate = yahoo_price("THB=X")
    if fx_rate:
        fx_ws["C6"] = round(fx_rate, 4)
        report["fx"] = round(fx_rate, 4)
    else:
        fx_rate = fx_old
        report["fx"] = fx_old
    if asof_date:
        fx_ws["C5"] = _dt.datetime(asof_date.year, asof_date.month, asof_date.day)
        fx_ws["D6"] = f"Auto-fetched via Yahoo Finance on {asof_date.isoformat()}"

    # ---- SGD/THB (for unit trusts) --------------------------------------
    if fx_ws["C9"].value is not None:                 # only if the SGD row exists
        sgd = None if offline else yahoo_price("SGDTHB=X")
        if sgd:
            fx_ws["C9"] = round(sgd, 4)
            report["sgd"] = round(sgd, 4)
            if asof_date:
                fx_ws["D9"] = f"Auto-fetched via Yahoo Finance on {asof_date.isoformat()}"

    # ---- collect crypto ids to batch ------------------------------------
    pws = wb["Prices"]
    # auto-managed daily-move reference columns (header on the row-4 label row)
    if pws.cell(4, 9).value != "Prev Price":
        pws.cell(4, 9).value = "Prev Price"
        pws.cell(4, 10).value = "Prev Date"
    crypto_ids = []
    for r in range(5, pws.max_row + 1):
        if pws.cell(r, 2).value == "Crypto":
            sym = (pws.cell(r, 3).value or "").upper()
            if sym in COINGECKO_IDS:
                crypto_ids.append(COINGECKO_IDS[sym])
    cg = {} if offline else coingecko_prices(crypto_ids)

    # ---- Thai mutual-fund NAVs via SEC Open API -------------------------
    mf_abbrs = [pws.cell(r, 3).value for r in range(5, pws.max_row + 1)
                if pws.cell(r, 2).value == "Mutual Fund" and pws.cell(r, 3).value]
    sec_navs = {}
    if not offline and mf_abbrs:
        sec_navs, report["sec"] = fetch_thai_nav.resolve_navs(mf_abbrs, asof_date)
    elif offline:
        report["sec"] = "offline"

    today_dt = _dt.datetime(asof_date.year, asof_date.month, asof_date.day) if asof_date else None

    # ---- walk the Prices sheet ------------------------------------------
    for r in range(5, pws.max_row + 1):
        typ = pws.cell(r, 2).value
        name = pws.cell(r, 3).value
        cust = pws.cell(r, 4).value
        ccy = (pws.cell(r, 6).value or "THB").upper()
        if not name:
            continue
        key = f"{typ}|{name}|{cust}"
        new_price = None
        prev_price = None          # prior trading day's price/NAV, same ccy as new_price
        prev_date = ""
        price_dt = today_dt   # date to stamp in the "Last Update" column

        if str(name).strip().lower() == "cash":
            pws.cell(r, 8).value = "OK"   # cash carries its balance; no market price
            continue
        if typ == "Unit Trust":
            continue                       # handled by the dedicated block below (symbol is in the Lots sheet)
        if offline:
            new_price = None
        elif typ == "Equity":
            new_price, prev_price = yahoo_quote(yahoo_symbol(name, ccy))
            time.sleep(0.15)
        elif typ == "Crypto":
            sym = name.upper()
            if sym in STABLECOINS:
                usd, usd_prev = STABLECOINS[sym], STABLECOINS[sym]
            else:
                hit = cg.get(COINGECKO_IDS.get(sym, ""))
                usd = hit["usd"] if hit else None
                usd_prev = hit.get("prev") if hit else None
            if usd is not None:
                new_price = usd if ccy == "USD" else usd * fx_rate
                if usd_prev is not None:
                    prev_price = usd_prev if ccy == "USD" else usd_prev * fx_rate
        elif typ == "Mutual Fund":
            hit = sec_navs.get(name)            # NAV from SEC Open API (THB)
            if hit:
                new_price = hit["nav"]
                prev_price = hit.get("prev")
                prev_date = hit.get("prev_date") or ""
                d = hit.get("date")
                if d:
                    try:
                        dd = _dt.date.fromisoformat(d)
                        price_dt = _dt.datetime(dd.year, dd.month, dd.day)
                    except ValueError:
                        pass

        if new_price and new_price > 0:
            pws.cell(r, 5).value = round(new_price, 6)
            if price_dt:
                pws.cell(r, 7).value = price_dt
            pws.cell(r, 8).value = "OK"
            if prev_price and prev_price > 0:    # daily-move reference (col I/J)
                pws.cell(r, 9).value = round(prev_price, 6)
                pws.cell(r, 10).value = prev_date
            report["updated"].append(key)
        else:
            # carry forward existing price; flag the source
            if typ == "Mutual Fund":
                report["carried"].append(key)
            elif not offline:
                report["failed"].append(key)
                pws.cell(r, 8).value = "MANUAL"

    # ---- SGD unit trusts: auto-fetch NAV into the Prices sheet -----------
    # Symbols live in the Lots sheet (col P); NAVs are written to the matching
    # "Unit Trust" rows in the Prices sheet (the Lots J column pulls from there).
    if "Unit Trust (SGD)" in wb.sheetnames and not offline:
        uws = wb["Unit Trust (SGD)"]
        price_row = {(pws.cell(r, 3).value, pws.cell(r, 4).value): r
                     for r in range(5, pws.max_row + 1) if pws.cell(r, 2).value == "Unit Trust"}
        r = 5
        while uws.cell(r, 1).value != "TOTAL" and r <= uws.max_row:
            fund = uws.cell(r, 3).value
            platform = uws.cell(r, 2).value
            symbol = uws.cell(r, 16).value          # P: Yahoo Symbol / ISIN
            if fund and symbol:
                sym = str(symbol).strip()
                px, px_prev = yahoo_quote(sym)
                if (not px) and re.fullmatch(r"[A-Za-z]{2}[A-Za-z0-9]{10}", sym):  # ISIN
                    resolved = yahoo_symbol_for_isin(sym)
                    if resolved:
                        px, px_prev = yahoo_quote(resolved)
                        if px and px > 0:
                            uws.cell(r, 16).value = resolved   # cache the Yahoo symbol
                time.sleep(0.15)
                prow = price_row.get((fund, platform))
                if px and px > 0 and prow:
                    pws.cell(prow, 5).value = round(px, 6)     # Prices NAV (SGD)
                    if today_dt:
                        pws.cell(prow, 7).value = today_dt
                    pws.cell(prow, 8).value = "OK"
                    if px_prev and px_prev > 0:                # daily-move reference
                        pws.cell(prow, 9).value = round(px_prev, 6)
                        pws.cell(prow, 10).value = ""
                    report["updated"].append(f"Unit Trust|{fund}")
                else:
                    if prow:
                        pws.cell(prow, 8).value = "MANUAL"
                    report["failed"].append(f"Unit Trust|{fund}")
            r += 1

    wb.save(workbook)
    return report


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--asof", default=_dt.date.today().isoformat())
    ap.add_argument("--workbook", default=str(WORKBOOK_DEFAULT))
    ap.add_argument("--offline", action="store_true", help="skip network; carry all prices forward")
    args = ap.parse_args()

    rep = fetch_all(args.workbook, args.asof, args.offline)
    print(f"FX USD/THB         : {rep['fx']}")
    print(f"Prices updated     : {len(rep['updated'])}")
    print(f"Carried (Thai MF)  : {len(rep['carried'])}")
    print(f"Failed -> MANUAL   : {len(rep['failed'])}")
    print(f"SEC NAV lookup     : {rep['sec']}")
    if rep["failed"]:
        print("   " + ", ".join(rep["failed"]))


if __name__ == "__main__":
    main()
