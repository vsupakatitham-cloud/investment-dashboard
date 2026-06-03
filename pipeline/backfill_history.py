"""
backfill_history.py — seed a daily portfolio history by assuming the CURRENT
holdings have been held since a start date (default 2026-05-01), priced with each
instrument's actual historical daily price and FX.

This is a "held-since" reconstruction (not actual realized performance — it
ignores past trades), used to give the Performance page a real daily return line
immediately. Each holding's daily THB value = current_value_thb x
(price[day]*fx[day]) / (price[asof]*fx[asof]); the latest day reconciles to the
real current total. Instruments without a fetchable history hold flat (FX still
varies for USD/SGD positions).

Sources: Yahoo (equities, crypto via COIN-USD, unit trusts via 0P..SI, USD/THB,
SGD/THB) and the SEC /v2/fund daily NAV API (Thai funds). Writes docs/history.json.

Usage:  SEC_OPENAPI_KEY=... python3 pipeline/backfill_history.py [--start 2026-05-01]
"""
from __future__ import annotations

import argparse
import datetime as _dt
import json
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl

import portfolio as _portfolio
import fetch_thai_nav as _ftn
from fetch_benchmark import _yahoo_daily
from fetch_prices import yahoo_symbol

ROOT = Path(__file__).resolve().parent.parent
DOCS = ROOT / "docs"
PIPE = Path(__file__).resolve().parent


def _safe_yahoo(sym, rng="6mo"):
    try:
        return _yahoo_daily(sym, rng)
    except Exception:
        return {}


def _sorted(d):
    return sorted(d.items()) if d else []


def _at(items, target_iso):
    """Forward-fill: value at the latest date <= target; clamp to the first."""
    if not items:
        return None
    lo, hi, ans = 0, len(items) - 1, items[0][1]
    if target_iso < items[0][0]:
        return items[0][1]
    while lo <= hi:
        mid = (lo + hi) // 2
        if items[mid][0] <= target_iso:
            ans = items[mid][1]
            lo = mid + 1
        else:
            hi = mid - 1
    return ans


def _sec_nav_series(proj_id, cls, key, start, end):
    out = {}
    cursor = None
    for _ in range(20):
        q = {"proj_id": proj_id, "start_nav_date": start, "end_nav_date": end, "page_size": 100}
        if cursor:
            q["next_cursor"] = cursor
        url = "https://api.sec.or.th/v2/fund/daily-info/nav?" + urllib.parse.urlencode(q)
        req = urllib.request.Request(url, headers={"Ocp-Apim-Subscription-Key": key,
                                    "Accept": "application/json", "Content-Type": "application/json"})
        try:
            d = json.loads(urllib.request.urlopen(req, timeout=30).read().decode())
        except Exception:
            break
        for it in d.get("items", []):
            if cls and it.get("fund_class_name") != cls:
                continue
            try:
                out[it["nav_date"]] = float(it["last_val"])
            except (TypeError, ValueError, KeyError):
                pass
        cursor = d.get("next_cursor")
        if not cursor:
            break
    return out


def build(start="2026-05-01", workbook=_portfolio.WORKBOOK_DEFAULT):
    p = _portfolio.load_portfolio(workbook)
    asof = p.as_of
    total_invested = round(p.total_invested)

    wb = openpyxl.load_workbook(workbook)
    ut_sym = {}
    if "Unit Trust (SGD)" in wb.sheetnames:
        uw = wb["Unit Trust (SGD)"]
        r = 5
        while uw.cell(r, 1).value != "TOTAL" and r < 60:
            if uw.cell(r, 3).value and uw.cell(r, 16).value:
                ut_sym[uw.cell(r, 3).value] = str(uw.cell(r, 16).value).strip()
            r += 1
    secmap = json.loads((PIPE / "sec_fund_map.json").read_text()).get("funds", {})
    key = _ftn._key()

    # FX histories
    usdthb = _sorted(_safe_yahoo("THB=X", "6mo"))
    sgdthb = _sorted(_safe_yahoo("SGDTHB=X", "6mo"))

    def fxmul(fxtype, d):
        if fxtype == "usd":
            return _at(usdthb, d) or 0
        if fxtype == "sgd":
            return _at(sgdthb, d) or 0
        return 1.0

    # pool holdings by (type,name); attach a price series + fx type
    pool = {}
    for h in p.holdings:
        k = (h["asset_type"], h["name"])
        g = pool.get(k)
        if not g:
            g = {"type": h["asset_type"], "name": h["name"],
                 "ccy": (h["currency"] or "THB").upper(), "value": 0.0}
            pool[k] = g
        g["value"] += h["value_thb"]

    print(f"Reconstructing {len(pool)} instruments from {start} to {asof} ...")
    for (t, name), g in pool.items():
        ccy = g["ccy"]
        price, fxtype = None, "thb"
        nm = str(name).strip()
        if nm.lower() == "cash":
            price, fxtype = None, ("usd" if ccy == "USD" else "thb")
        elif t == "Equity":
            price = _sorted(_safe_yahoo(yahoo_symbol(nm, ccy), "6mo"))
            fxtype = "thb" if ccy == "THB" else "usd"
        elif t == "Crypto":
            if nm.upper() == "USDT":
                price, fxtype = None, "usd"
            else:
                price = _sorted(_safe_yahoo(nm.upper() + "-USD", "6mo"))
                fxtype = "usd"
        elif t == "MF":
            pin = secmap.get(nm)
            if pin and key:
                price = _sorted(_sec_nav_series(pin["proj_id"], pin.get("fund_class_name"), key, start, asof))
            fxtype = "thb"
        elif t == "Unit Trust":
            sym = ut_sym.get(nm)
            if sym:
                price = _sorted(_safe_yahoo(sym, "6mo"))
            fxtype = "sgd"
        g["price"], g["fxtype"] = price, fxtype
        g["driver_asof"] = (_at(price, asof) if price else 1.0) * fxmul(fxtype, asof)
        if not g["driver_asof"]:
            g["driver_asof"] = 1.0
            g["price"] = None  # fall back to flat

    # iterate calendar days
    d0 = _dt.date.fromisoformat(start)
    d1 = _dt.date.fromisoformat(asof)
    rows = []
    day = d0
    while day <= d1:
        diso = day.isoformat()
        tot = 0.0
        sleeve = {"MF": 0.0, "Equity": 0.0, "Crypto": 0.0}
        for g in pool.values():
            driver = (_at(g["price"], diso) if g["price"] else 1.0) * fxmul(g["fxtype"], diso)
            val = g["value"] * (driver / g["driver_asof"])
            tot += val
            if g["type"] in sleeve:
                sleeve[g["type"]] += val
        pnl = tot - total_invested
        rows.append({
            "date": diso, "total_value": round(tot), "total_invested": total_invested,
            "total_pnl": round(pnl), "pnl_pct": round(pnl / total_invested, 4) if total_invested else 0,
            "mf_value": round(sleeve["MF"]), "eq_value": round(sleeve["Equity"]),
            "crypto_value": round(sleeve["Crypto"]), "fx_rate": p.fx_rate,
        })
        day += _dt.timedelta(days=1)

    DOCS.mkdir(exist_ok=True)
    (DOCS / "history.json").write_text(json.dumps(rows, indent=2))
    return rows


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--start", default="2026-05-01")
    args = ap.parse_args()
    rows = build(args.start)
    print(f"Wrote {len(rows)} daily points -> docs/history.json")
    print(f"  {rows[0]['date']}: ฿{rows[0]['total_value']:,}  ({rows[0]['pnl_pct']:+.2%})")
    print(f"  {rows[-1]['date']}: ฿{rows[-1]['total_value']:,}  ({rows[-1]['pnl_pct']:+.2%})")
