"""
run_weekly.py — One-shot weekly update for the private-client dashboard.

Pipeline:
  1. Fetch live prices into the workbook  (Yahoo + CoinGecko; Thai MF carried fwd)
  2. Append a row to the workbook's "Weekly Snapshot" history log
  3. Rebuild docs/ (data.json, history.json, index.html)
  4. Commit & push to GitHub Pages  (unless --no-publish)

Intended to be run every Saturday 09:00 Asia/Bangkok by the scheduler.

Usage:
    python3 pipeline/run_weekly.py                 # full run + publish
    python3 pipeline/run_weekly.py --no-publish    # local only
    python3 pipeline/run_weekly.py --offline       # no network price fetch
"""
from __future__ import annotations

import argparse
import datetime as _dt
import subprocess
import sys
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "pipeline"))

from fetch_prices import fetch_all, WORKBOOK_DEFAULT          # noqa: E402
from portfolio import load_portfolio                          # noqa: E402
from build_site import build                                  # noqa: E402

BKK = ZoneInfo("Asia/Bangkok")


def log_snapshot(workbook: Path, p) -> None:
    """Append the current week's totals to the Weekly Snapshot sheet (history log)."""
    wb = openpyxl.load_workbook(workbook, data_only=False)
    ws = wb["Weekly Snapshot"]
    asof = _dt.date.fromisoformat(p.as_of)
    asof_dt = _dt.datetime(asof.year, asof.month, asof.day)
    # Header row is 4; live row is 5; history from row 6 downward.
    # Reuse a row already logged for this as-of date (idempotent re-runs).
    target = 6
    while ws.cell(target, 2).value not in (None, ""):
        existing = ws.cell(target, 2).value
        if existing == asof_dt or (isinstance(existing, _dt.datetime) and existing.date() == asof):
            break
        target += 1
    vals = [
        _dt.datetime(asof.year, asof.month, asof.day),  # B Date
        round(p.total_value), round(p.total_invested), round(p.total_pnl),
        round(p.total_pnl_pct, 4), round(p.mf_value), round(p.eq_value),
        round(p.crypto_value), p.fx_rate, "Auto weekly snapshot",
    ]
    for i, v in enumerate(vals):
        ws.cell(target, 2 + i).value = v
    wb.save(workbook)
    print(f"  snapshot logged to Weekly Snapshot row {target}")


def git_publish(asof: str) -> bool:
    def run(*a):
        return subprocess.run(["git", *a], cwd=ROOT, capture_output=True, text=True)
    if not (ROOT / ".git").exists():
        print("  [publish] no git repo yet — run scripts/setup_github.sh first. Skipping push.")
        return False
    run("add", "docs", "TH Investment - Private Banking Summary.xlsx")
    status = run("status", "--porcelain")
    if not status.stdout.strip():
        print("  [publish] nothing changed.")
        return True
    run("commit", "-m", f"Weekly dashboard update — {asof}")
    push = run("push", "origin", "HEAD")
    if push.returncode == 0:
        print("  [publish] pushed to GitHub Pages.")
        return True
    print("  [publish] push failed:\n" + push.stderr)
    return False


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", default=str(WORKBOOK_DEFAULT))
    ap.add_argument("--asof", default=None, help="override as-of date (YYYY-MM-DD)")
    ap.add_argument("--offline", action="store_true")
    ap.add_argument("--no-publish", action="store_true")
    args = ap.parse_args()

    now = _dt.datetime.now(BKK)
    asof = args.asof or now.date().isoformat()
    workbook = Path(args.workbook)
    print(f"Weekly update — as of {asof}  ({now:%Y-%m-%d %H:%M %Z})")

    print("[1/4] Fetching live prices ...")
    rep = fetch_all(workbook, asof, offline=args.offline)
    print(f"      FX {rep['fx']} · {len(rep['updated'])} updated · "
          f"{len(rep['carried'])} MF carried · {len(rep['failed'])} manual")
    print(f"      SEC NAV: {rep.get('sec','')}")

    print("[2/4] Logging weekly snapshot ...")
    p = load_portfolio(workbook)
    log_snapshot(workbook, p)

    print("[3/4] Building dashboard ...")
    if not args.offline:
        try:
            import fetch_benchmark
            b = fetch_benchmark.build(asof)
            if b:
                print(f"      benchmark {b['name']}: {len(b['series'])} pts")
        except Exception as e:
            print(f"      benchmark fetch skipped ({e})")
    out = build(workbook, generated_at=now.strftime("%Y-%m-%d %H:%M %Z"))
    pf = out["portfolio"]
    perf = out.get("performance", {})
    print(f"      Total ฿{pf['total_value']:,.0f} · P&L {pf['total_pnl_pct']:+.2%} · "
          f"{len(out['history'])} history point(s) · perf history {perf.get('history_points','?')}d")

    print("[4/4] Publishing ...")
    if args.no_publish:
        print("      skipped (--no-publish)")
    else:
        git_publish(asof)

    print("Done.")


if __name__ == "__main__":
    main()
