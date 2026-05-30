"""
make_template.py — Produce a blank, reusable onboarding template from the
working workbook. Every formula, named range, layout and color rule is kept;
only the client's positions and prices are cleared so the file can be dropped
onto a new private client.

Output: template/TH Investment - Private Banking TEMPLATE.xlsx

Input columns cleared per sheet (formula columns are left untouched so the
workbook stays fully live the moment new lots are typed in):
  MF - Lots          B,C,H,I,J,L,M,O,S
  Equities - Lots    B,C,D,H,I,P
  Crypto - Lots      B,C,D,F,G,N,O
  Prices             B,C,D,E,F,G,H
  Reference          B..J
  MF - by Fund       B,C        (rollup name lists)
  Equities-by-Ticker B,C,D
  Crypto - by Coin   B,C,D
  Weekly Snapshot    rows 6+    (history log)
  FX & Assumptions   reset date / rate / counts to placeholders
"""
from __future__ import annotations

import datetime as _dt
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "TH Investment - Private Banking Summary.xlsx"
OUT = ROOT / "template" / "TH Investment - Private Banking TEMPLATE.xlsx"

CLEAR = {
    "MF - Lots":            (5, 164, [2, 3, 8, 9, 10, 12, 13, 15, 19]),
    "Equities - Lots":      (5, 45,  [2, 3, 4, 8, 9, 16]),
    "Crypto - Lots":        (5, 30,  [2, 3, 4, 6, 7, 14, 15]),
    "Prices":               (5, 96,  [2, 3, 4, 5, 6, 7, 8]),
    "Reference":            (5, 99,  [2, 3, 4, 5, 6, 7, 8, 9, 10]),
    "MF - by Fund":         (5, 36,  [2, 3]),
    "Equities - by Ticker": (5, 43,  [2, 3, 4]),
    "Crypto - by Coin":     (5, 28,  [2, 3, 4]),
}


def main():
    OUT.parent.mkdir(exist_ok=True)
    wb = openpyxl.load_workbook(SRC, data_only=False)

    for sheet, (first, last, cols) in CLEAR.items():
        ws = wb[sheet]
        for r in range(first, last + 1):
            for c in cols:
                ws.cell(r, c).value = None

    # Weekly Snapshot: clear appended history (row 6 down)
    ws = wb["Weekly Snapshot"]
    for r in range(6, ws.max_row + 1):
        for c in range(2, 12):
            ws.cell(r, c).value = None

    # FX & Assumptions: reset to placeholders
    fx = wb["FX & Assumptions"]
    fx["C5"] = _dt.datetime(_dt.date.today().year, 1, 1)
    fx["C6"] = 33.0
    fx["D6"] = "Enter weekly USD/THB spot (or let the auto-fetch fill it)."
    fx["C8"] = "0 / 0 / 0"

    wb.save(OUT)
    print(f"Blank template written: {OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
