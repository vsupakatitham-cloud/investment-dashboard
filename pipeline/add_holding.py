"""
add_holding.py — Add a new holding to the workbook (and wire up pricing) so the
dashboard picks it up. Writes the three places a holding needs:

  1. the matching "… - Lots" sheet  (a new lot row above TOTAL, formulas copied)
  2. the Reference sheet            (classification, keyed Type|Name|Custodian)
  3. the Prices sheet               (so the weekly job can value it)

and the pricing hookup:
  * equity  → priced automatically via Yahoo (Thai tickers get ".BK")
  * crypto  → records a CoinGecko id in pipeline/coingecko_ids.json (else manual)
  * mutual fund → pins the SEC proj_id + exact share class in sec_fund_map.json
                  (lists the available classes so you confirm the right one)

The workbook is backed up first, and the result is validated by reloading it.

Examples
--------
  python3 pipeline/add_holding.py equity --ticker GOOGL --broker Dime --ccy USD \
      --shares 10 --avg-cost 150 --asset-class Equity --geography "United States" \
      --theme "US Large Cap" --tax Taxable

  python3 pipeline/add_holding.py crypto --coin XRP --exchange Binance --ccy USD \
      --qty 100 --avg-cost 0.50 --subclass "Alt L1" --coingecko-id ripple

  python3 pipeline/add_holding.py mf --fund K-GOLD --amc "K Asset" --units 1000 \
      --avg-cost 12 --initial 12000 --date 2026-05-31 --asset-class Alternatives \
      --geography Global --theme Gold --tax Open-Ended           # then pick the class

Run with no flags for an interactive prompt.
"""
from __future__ import annotations

import argparse
import datetime as _dt
import json
import re
import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.formula.translate import Translator

ROOT = Path(__file__).resolve().parent.parent
WB = ROOT / "TH Investment - Private Banking Summary.xlsx"
PIPE = Path(__file__).resolve().parent

# Per-asset layout of each Lots sheet: input columns (name->col idx), formula
# columns to copy from the row above, and the TOTAL-row formula templates.
SPECS = {
    "equity": dict(
        sheet="Equities - Lots", wb_type="Equity",
        inputs=dict(broker=2, ticker=3, ccy=4, shares=8, avg_cost=9, note=16),
        formula_cols=[5, 6, 7, 10, 11, 12, 13, 14, 15],
        totals={12: "=SUM(L5:L{last})", 13: "=SUM(M5:M{last})", 14: "=SUM(N5:N{last})",
                15: "=IFERROR(N{tr}/L{tr},0)"},
        name="ticker", custodian="broker",
    ),
    "crypto": dict(
        sheet="Crypto - Lots", wb_type="Crypto",
        inputs=dict(exchange=2, coin=3, ccy=4, qty=6, avg_cost=7, date=14, note=15),
        formula_cols=[5, 8, 9, 10, 11, 12, 13],
        totals={10: "=SUM(J5:J{last})", 11: "=SUM(K5:K{last})", 12: "=SUM(L5:L{last})",
                13: "=IFERROR(L{tr}/J{tr},0)"},
        name="coin", custodian="exchange",
    ),
    "mf": dict(
        sheet="MF - Lots", wb_type="Mutual Fund",
        inputs=dict(amc=2, fund=3, date=8, year=9, sellable=10, units=12, avg_cost=13,
                    initial=15, note=19),
        formula_cols=[4, 5, 6, 7, 11, 14, 16, 17, 18],
        totals={15: "=SUM(O5:O{last})", 16: "=SUM(P5:P{last})", 17: "=SUM(Q5:Q{last})",
                18: "=IFERROR(Q{tr}/O{tr},0)"},
        name="fund", custodian="amc",
    ),
}


def _totals_row(ws, name_col=3):
    r = 5
    while r <= ws.max_row:
        if isinstance(ws.cell(r, 1).value, str) and ws.cell(r, 1).value.strip().upper() == "TOTAL":
            return r
        r += 1
    # no TOTAL row found — append after the last named row
    last = 5
    for rr in range(5, ws.max_row + 1):
        if ws.cell(rr, name_col).value not in (None, ""):
            last = rr
    return last + 1


def _col(idx):
    return openpyxl.utils.get_column_letter(idx)


def insert_lot(wb, spec, values):
    ws = wb[spec["sheet"]]
    T = _totals_row(ws)             # TOTAL row index (new lot goes here)
    TEMPLATE = 5                    # first data row holds the canonical formulas
    # snapshot the template row's formulas BEFORE inserting (avoids copying an
    # atypical neighbour like a hard-coded "Cash" row)
    src_formulas = {c: ws.cell(TEMPLATE, c).value for c in spec["formula_cols"]}
    ws.insert_rows(T)              # blank row at T; TOTAL shifts to T+1

    # write inputs
    for key, cidx in spec["inputs"].items():
        if key in values and values[key] is not None:
            ws.cell(T, cidx).value = values[key]
    ws.cell(T, 1).value = T - 4    # running number (#); row5 == #1
    # copy canonical formulas, translated to the new row
    for c, f in src_formulas.items():
        if isinstance(f, str) and f.startswith("="):
            ws.cell(T, c).value = Translator(f, origin=f"{_col(c)}{TEMPLATE}").translate_formula(f"{_col(c)}{T}")

    # rebuild TOTAL row (now at T+1) so its ranges include the new row T
    tr = T + 1
    ws.cell(tr, 1).value = "TOTAL"
    for cidx, tmpl in spec["totals"].items():
        ws.cell(tr, cidx).value = tmpl.format(last=T, tr=tr)

    # bump every cross-sheet reference that pointed at the old TOTAL row T -> T+1
    pat = re.compile(r"('%s'!\$?)([A-Z]{1,2})(\$?)%d(?![0-9])" % (re.escape(spec["sheet"]), T))
    for w in wb.worksheets:
        for row in w.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("=") and pat.search(v):
                    cell.value = pat.sub(lambda m: f"{m.group(1)}{m.group(2)}{m.group(3)}{T+1}", v)
    return T


def _first_blank(ws, name_col=3, first=5):
    r = first
    while ws.cell(r, name_col).value not in (None, ""):
        r += 1
    return r


def ensure_reference(wb, wb_type, name, cust, meta):
    ws = wb["Reference"]
    key = f"{wb_type}|{name}|{cust}"
    for r in range(5, ws.max_row + 1):
        if ws.cell(r, 2).value == wb_type and ws.cell(r, 3).value == name and ws.cell(r, 4).value == cust:
            return f"Reference already had {key}"
    r = _first_blank(ws)
    ws.cell(r, 1).value = f'=B{r}&"|"&C{r}&"|"&D{r}'
    ws.cell(r, 2).value = wb_type
    ws.cell(r, 3).value = name
    ws.cell(r, 4).value = cust
    ws.cell(r, 5).value = meta.get("asset_class", "")
    ws.cell(r, 6).value = meta.get("sub_class", "")
    ws.cell(r, 7).value = meta.get("geography", "")
    ws.cell(r, 8).value = meta.get("theme", "")
    ws.cell(r, 9).value = meta.get("tax_status", "")
    ws.cell(r, 10).value = meta.get("currency", "THB")
    ws.cell(r, 11).value = f'=IF(J{r}="THB",1,FXRate)'
    return f"Reference row added at {r}"


def ensure_price(wb, wb_type, name, cust, ccy, price=None, check="MANUAL"):
    ws = wb["Prices"]
    for r in range(5, ws.max_row + 1):
        if ws.cell(r, 2).value == wb_type and ws.cell(r, 3).value == name and ws.cell(r, 4).value == cust:
            return f"Prices already had {wb_type}|{name}|{cust}"
    r = _first_blank(ws)
    ws.cell(r, 1).value = f'=B{r}&"|"&C{r}&"|"&D{r}'
    ws.cell(r, 2).value = wb_type
    ws.cell(r, 3).value = name
    ws.cell(r, 4).value = cust
    ws.cell(r, 5).value = price
    ws.cell(r, 6).value = ccy
    ws.cell(r, 8).value = check
    return f"Prices row added at {r}"


def add_coingecko_id(symbol, cg_id):
    f = PIPE / "coingecko_ids.json"
    d = json.loads(f.read_text()) if f.exists() else {}
    d[symbol.upper()] = cg_id
    f.write_text(json.dumps(d, indent=1))
    return f"Recorded CoinGecko id {symbol.upper()} -> {cg_id}"


def pin_mf_class(fund, chosen_class=None, asof=None):
    """Resolve a fund's SEC proj_id + share class and write it into sec_fund_map.json.
    Returns (message, needs_choice_list_or_None)."""
    sys.path.insert(0, str(PIPE))
    try:
        import fetch_thai_nav as ftn
    except Exception as e:
        return f"(SEC module unavailable: {e})", None
    key = ftn._key()
    if not key:
        return "SEC key not set — fund will carry forward until you pin it (set SEC_OPENAPI_KEY).", None
    rows = ftn._all_profiles(key)
    norm = lambda s: re.sub(r"\s+", "", (s or "")).upper()
    # candidates: exact class name, exact abbr, else loose contains
    cands = {}
    for it in rows:
        ab, cls, pid = it.get("proj_abbr_name"), it.get("fund_class_name"), it.get("proj_id")
        if norm(cls) == norm(fund) or norm(ab) == norm(fund):
            cands[(pid, cls)] = True
    if not cands:
        for it in rows:
            ab, cls, pid = it.get("proj_abbr_name"), it.get("fund_class_name"), it.get("proj_id")
            if norm(fund) in norm(ab) or norm(fund) in norm(cls):
                cands[(pid, cls)] = True
    cands = list(cands)
    if not cands:
        return f"No SEC match for {fund} — pin it by hand in sec_fund_map.json.", None
    if chosen_class:
        for pid, cls in cands:
            if norm(cls) == norm(chosen_class):
                _write_pin(fund, pid, cls)
                return f"Pinned {fund} -> {pid} / {cls}", None
        return f"--class {chosen_class} not found among {[c for _,c in cands]}", None
    if len(cands) == 1:
        pid, cls = cands[0]
        _write_pin(fund, pid, cls)
        return f"Pinned {fund} -> {pid} / {cls} (single class)", None
    # multiple classes: ask the user to pick, with latest NAVs to help
    asof = asof or _dt.date.today()
    listing = []
    for pid, cls in cands:
        nav = ftn.latest_nav(pid, cls, key, asof)
        listing.append((pid, cls, nav[0] if nav else None, nav[1] if nav else ""))
    return f"{fund} has multiple share classes — re-run with --class <name>:", listing


def _write_pin(fund, pid, cls):
    f = PIPE / "sec_fund_map.json"
    d = json.loads(f.read_text()) if f.exists() else {"funds": {}}
    d.setdefault("funds", {})[fund] = {"proj_id": pid, "fund_class_name": cls}
    f.write_text(json.dumps(d, indent=1, ensure_ascii=False))


def run(asset, values, meta, *, coingecko_id=None, mf_class=None, workbook=WB, dry=False):
    workbook = Path(workbook)
    spec = SPECS[asset]
    name, cust = values[spec["name"]], values[spec["custodian"]]
    ccy = values.get("ccy", "THB")

    # For funds, resolve the SEC share class FIRST. If the fund has several
    # classes and none was chosen, stop here and list them — nothing is written,
    # so re-running with --class won't create a duplicate lot.
    mf_msg, pending_classes = None, None
    if asset == "mf":
        mf_msg, pending_classes = pin_mf_class(name, mf_class)
        if pending_classes:
            return [mf_msg], pending_classes

    backup = workbook.with_suffix(".prechange.xlsx")
    shutil.copy(workbook, backup)
    wb = openpyxl.load_workbook(workbook, data_only=False)

    log = []
    T = insert_lot(wb, spec, values)
    log.append(f"{spec['sheet']}: lot added at row {T}")
    log.append(ensure_reference(wb, spec["wb_type"], name, cust, meta))

    # pricing hookup
    check, price = "OK", None
    if asset == "crypto" and coingecko_id:
        log.append(add_coingecko_id(name, coingecko_id))
    elif asset == "crypto":
        check = "MANUAL"
    if asset == "mf":
        log.append(mf_msg)
    log.append(ensure_price(wb, spec["wb_type"], name, cust, ccy, price, check))

    if dry:
        backup.unlink(missing_ok=True)
        return log, pending_classes
    wb.save(workbook)

    # validate: reload via the engine
    try:
        sys.path.insert(0, str(PIPE))
        import importlib, portfolio
        importlib.reload(portfolio)
        p = portfolio.load_portfolio(workbook)
        log.append(f"validated: {sum(p.lot_counts.values())} lots, total ฿{p.total_value:,.0f}")
        backup.unlink(missing_ok=True)
    except Exception as e:
        shutil.copy(backup, workbook)
        log.append(f"ERROR during validation — restored backup ({e})")
    return log, pending_classes


def _prompt(args, asset):
    spec = SPECS[asset]
    need = list(spec["inputs"]) + ["asset_class", "geography", "theme", "tax"]
    for k in need:
        if getattr(args, k.replace("-", "_"), None) is None and k not in ("note", "year", "sellable", "date"):
            if sys.stdin.isatty():
                v = input(f"  {k.replace('_',' ')}: ").strip()
                setattr(args, k.replace("-", "_"), v or None)


def main():
    ap = argparse.ArgumentParser(description="Add a holding to the dashboard workbook.")
    sub = ap.add_subparsers(dest="asset", required=True)
    common = dict()
    for a in ("equity", "crypto", "mf"):
        p = sub.add_parser(a)
        p.add_argument("--asset-class"); p.add_argument("--geography")
        p.add_argument("--theme"); p.add_argument("--tax"); p.add_argument("--subclass")
        p.add_argument("--ccy", default="THB"); p.add_argument("--note")
        p.add_argument("--workbook", default=str(WB))
        if a == "equity":
            p.add_argument("--ticker"); p.add_argument("--broker")
            p.add_argument("--shares", type=float); p.add_argument("--avg-cost", type=float)
        elif a == "crypto":
            p.add_argument("--coin"); p.add_argument("--exchange")
            p.add_argument("--qty", type=float); p.add_argument("--avg-cost", type=float)
            p.add_argument("--coingecko-id"); p.add_argument("--date")
        else:
            p.add_argument("--fund"); p.add_argument("--amc")
            p.add_argument("--units", type=float); p.add_argument("--avg-cost", type=float)
            p.add_argument("--initial", type=float); p.add_argument("--date")
            p.add_argument("--year", type=int); p.add_argument("--sellable")
            p.add_argument("--class", dest="mf_class")
    args = ap.parse_args()
    asset = args.asset
    spec = SPECS[asset]
    _prompt(args, asset)

    values = {}
    for key in spec["inputs"]:
        values[key] = getattr(args, key, None)
    values["ccy"] = args.ccy
    meta = dict(asset_class=args.asset_class, sub_class=getattr(args, "subclass", "") or "",
                geography=args.geography, theme=args.theme, tax_status=args.tax, currency=args.ccy)

    missing = [spec["name"], spec["custodian"]] + (["shares" if asset == "equity" else "qty" if asset == "crypto" else "units"])
    for m in missing:
        if not values.get(m):
            ap.error(f"missing required: --{m}")

    log, pending = run(asset, values, meta,
                       coingecko_id=getattr(args, "coingecko_id", None),
                       mf_class=getattr(args, "mf_class", None),
                       workbook=args.workbook)
    print("\n".join("  • " + l for l in log))
    if pending:
        print("\n  Share classes for this fund (re-run with --class <name>):")
        for pid, cls, nav, d in pending:
            print(f"    --class {cls:<18} proj {pid}  NAV {nav if nav else '—'} {d}")
        return
    print("\nDone. Rebuild with:  python3 pipeline/run_weekly.py --no-publish")


if __name__ == "__main__":
    main()
